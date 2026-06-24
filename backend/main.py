import io
import os
import sys
import json
import sqlite3
import shutil
import pandas as pd
import xlsxwriter
from datetime import datetime, date
from typing import List, Optional
from pydantic import BaseModel
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

# ReportLab imports
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

# Setup persistent directory in User's Documents to keep Desktop clean
doc_dir = os.path.join(os.path.expanduser("~"), "Documents", "MetroRailwayERP")
os.makedirs(doc_dir, exist_ok=True)
DB_PATH = os.getenv("ERP_DB_PATH", os.path.join(doc_dir, "database.db"))
BACKUP_DIR = os.path.join(doc_dir, "backups")

# Ensure directories exist
os.makedirs(BACKUP_DIR, exist_ok=True)

# Copy default database from PyInstaller bundle if it doesn't exist
if not os.path.exists(DB_PATH):
    if getattr(sys, 'frozen', False):
        default_db = os.path.join(sys._MEIPASS, "database.db")
    else:
        default_db = os.path.join(os.path.dirname(os.path.abspath(__file__)), "database.db")
        
    if os.path.exists(default_db):
        try:
            shutil.copy(default_db, DB_PATH)
        except Exception as e:
            print(f"Failed to copy default database: {e}")

TA_TEMPLATE_PATH = os.path.join(doc_dir, "TA bill.xlsx")
if not os.path.exists(TA_TEMPLATE_PATH):
    if getattr(sys, 'frozen', False):
        default_ta = os.path.join(sys._MEIPASS, "TA bill.xlsx")
    else:
        default_ta = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "TA bill.xlsx")
    if os.path.exists(default_ta):
        try:
            shutil.copy(default_ta, TA_TEMPLATE_PATH)
        except Exception as e:
            print(f"Failed to copy default TA template: {e}")

app = FastAPI(title="Metro Railway Kolkata S&T Staff Management System Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- SQLite Database Helper ---
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.row_factory = sqlite3.Row
    return conn

def get_section_name_from_db(section_code: str) -> str:
    if not section_code:
        return ""
    try:
        conn = get_db()
        row = conn.execute("SELECT section_name FROM sections WHERE section_code = ?", (section_code,)).fetchone()
        conn.close()
        return row['section_name'] if row else section_code
    except Exception as e:
        print(f"Error fetching section name: {e}")
        return section_code


# --- Audit Logging Utility ---
def log_audit(action: str, module: str, details: str, user: str = "System Admin"):
    try:
        conn = get_db()
        conn.execute("""
            INSERT INTO audit_logs (timestamp, user, module, action, details)
            VALUES (?, ?, ?, ?, ?)
        """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), user, module, action, details))
        conn.commit()
        conn.close()
    except Exception as e:
        print("Audit log failed:", e)

# --- Recalculate Leave Bank & Sync CR Ledger ---
def sync_leave_and_ledger(emp_id: int, year: int):
    conn = get_db()
    cursor = conn.cursor()
    try:
        # Get employee rest day
        cursor.execute("SELECT default_rest_day FROM employees WHERE emp_id = ?", (emp_id,))
        emp = cursor.fetchone()
        if not emp:
            return
        rest_day = emp['default_rest_day']

        start_date = f"{year}-01-01"
        end_date = f"{year}-12-31"

        # Get all logs for employee in that year
        cursor.execute("""
            SELECT date, status, remarks FROM attendance_log 
            WHERE emp_id = ? AND date >= ? AND date <= ?
        """, (emp_id, start_date, end_date))
        logs = cursor.fetchall()

        used_cl = sum(1 for log in logs if log['status'] == 'CL')
        used_lap = sum(1 for log in logs if log['status'] == 'LAP')

        # Earned CRs (Working P or P/N on their rest day)
        earned_dates = []
        for log in logs:
            if log['status'] in ['P', 'P/N']:
                dt = datetime.strptime(log['date'], "%Y-%m-%d")
                day_name = dt.strftime("%A")
                if day_name == rest_day:
                    earned_dates.append(log['date'])

        # Gather explicitly mapped CRs and unassociated CRs
        consumed_dates = [log['date'] for log in logs if log['status'] == 'CR']
        explicit_mappings = {} # consumed_date -> earned_date
        unassociated_consumed = []
        for log in logs:
            if log['status'] == 'CR':
                cdate = log['date']
                remarks = log['remarks'] or ''
                if "CR_EARNED_DATE:" in remarks:
                    try:
                        parts = remarks.split("CR_EARNED_DATE:")
                        if len(parts) > 1:
                            edate = parts[1].strip()[:10]
                            datetime.strptime(edate, "%Y-%m-%d")
                            explicit_mappings[cdate] = edate
                            continue
                    except Exception:
                        pass
                unassociated_consumed.append(cdate)

        # Update Compensatory Rest Ledger
        # Fetch current ledger
        cursor.execute("SELECT id, earned_date, consumed_date FROM compensatory_rest_ledger WHERE emp_id = ?", (emp_id,))
        ledger = {row['earned_date']: dict(row) for row in cursor.fetchall()}

        # Insert missing earned dates
        for edate in earned_dates:
            if edate not in ledger:
                cursor.execute("INSERT INTO compensatory_rest_ledger (emp_id, earned_date, consumed_date) VALUES (?, ?, NULL)", (emp_id, edate))
                ledger[edate] = {'earned_date': edate, 'consumed_date': None}

        # Delete invalid earned dates
        for edate in list(ledger.keys()):
            if edate not in earned_dates:
                cursor.execute("DELETE FROM compensatory_rest_ledger WHERE emp_id = ? AND earned_date = ?", (emp_id, edate))
                del ledger[edate]

        # Reset all consumed dates to NULL first
        cursor.execute("UPDATE compensatory_rest_ledger SET consumed_date = NULL WHERE emp_id = ?", (emp_id,))

        # Apply explicit mappings
        paired_earned = set()
        for cdate, edate in explicit_mappings.items():
            cursor.execute("""
                INSERT INTO compensatory_rest_ledger (emp_id, earned_date, consumed_date)
                VALUES (?, ?, ?)
                ON CONFLICT(emp_id, earned_date) DO UPDATE SET consumed_date = excluded.consumed_date
            """, (emp_id, edate, cdate))
            paired_earned.add(edate)

        # Pair remaining unassociated consumed CRs with remaining unconsumed earned CRs chronologically
        remaining_earned = [edate for edate in earned_dates if edate not in paired_earned]
        remaining_earned.sort()
        unassociated_consumed.sort()

        c_idx = 0
        for edate in remaining_earned:
            if c_idx < len(unassociated_consumed):
                cursor.execute("""
                    UPDATE compensatory_rest_ledger 
                    SET consumed_date = ? 
                    WHERE emp_id = ? AND earned_date = ?
                """, (unassociated_consumed[c_idx], emp_id, edate))
                c_idx += 1

        # Save accrued_cr count
        accrued_cr = max(0, len(earned_dates) - len(consumed_dates))

        # Upsert leave bank
        cursor.execute("""
            INSERT INTO leave_bank (emp_id, year, total_cl, total_lap, used_cl, used_lap, accrued_cr)
            VALUES (?, ?, 8, 30, ?, ?, ?)
            ON CONFLICT(emp_id, year) DO UPDATE SET
                used_cl = excluded.used_cl,
                used_lap = excluded.used_lap,
                accrued_cr = excluded.accrued_cr
        """, (emp_id, year, used_cl, used_lap, accrued_cr))

        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"Failed to sync leave & ledger for employee {emp_id}:", e)
    finally:
        conn.close()

# --- Database Initialization & Seeding ---
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Rebuild tables referencing employees_old if any are corrupted by SQLite RENAME
    for table_name in ('compensatory_rest_ledger', 'leave_bank', 'attendance_log', 'special_events'):
        cursor.execute(f"SELECT sql FROM sqlite_master WHERE type='table' AND name='{table_name}'")
        row = cursor.fetchone()
        if row and 'employees_old' in row[0]:
            print(f"Repairing table {table_name} to reference new employees table...")
            try:
                cursor.execute("PRAGMA foreign_keys = OFF;")
                cursor.execute(f"ALTER TABLE {table_name} RENAME TO {table_name}_old;")
                
                if table_name == 'compensatory_rest_ledger':
                    cursor.execute("""
                        CREATE TABLE compensatory_rest_ledger (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            emp_id INTEGER REFERENCES employees(emp_id) ON DELETE CASCADE,
                            earned_date TEXT NOT NULL,
                            consumed_date TEXT,
                            UNIQUE (emp_id, earned_date)
                        );
                    """)
                    cursor.execute("""
                        INSERT INTO compensatory_rest_ledger (id, emp_id, earned_date, consumed_date)
                        SELECT id, emp_id, earned_date, consumed_date FROM compensatory_rest_ledger_old;
                    """)
                elif table_name == 'leave_bank':
                    cursor.execute("""
                        CREATE TABLE leave_bank (
                            emp_id INTEGER REFERENCES employees(emp_id) ON DELETE CASCADE,
                            year INTEGER NOT NULL,
                            total_cl INTEGER NOT NULL DEFAULT 8,
                            total_lap INTEGER NOT NULL DEFAULT 30,
                            used_cl INTEGER NOT NULL DEFAULT 0,
                            used_lap INTEGER NOT NULL DEFAULT 0,
                            accrued_cr INTEGER NOT NULL DEFAULT 0,
                            PRIMARY KEY (emp_id, year)
                        );
                    """)
                    cursor.execute("""
                        INSERT INTO leave_bank (emp_id, year, total_cl, total_lap, used_cl, used_lap, accrued_cr)
                        SELECT emp_id, year, total_cl, total_lap, used_cl, used_lap, accrued_cr FROM leave_bank_old;
                    """)
                elif table_name == 'attendance_log':
                    cursor.execute("""
                        CREATE TABLE attendance_log (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            emp_id INTEGER REFERENCES employees(emp_id) ON DELETE CASCADE,
                            date TEXT NOT NULL,
                            status TEXT NOT NULL,
                            is_night INTEGER NOT NULL DEFAULT 0,
                            shift_id INTEGER REFERENCES shift_rules(id) ON DELETE SET NULL,
                            remarks TEXT,
                            UNIQUE (emp_id, date)
                        );
                    """)
                    cursor.execute("""
                        INSERT INTO attendance_log (id, emp_id, date, status, is_night, shift_id, remarks)
                        SELECT id, emp_id, date, status, is_night, shift_id, remarks FROM attendance_log_old;
                    """)
                elif table_name == 'special_events':
                    cursor.execute("""
                        CREATE TABLE special_events (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            emp_id INTEGER REFERENCES employees(emp_id) ON DELETE CASCADE,
                            event_type TEXT NOT NULL,
                            from_date TEXT NOT NULL,
                            to_date TEXT NOT NULL,
                            order_number TEXT,
                            location TEXT,
                            from_section TEXT,
                            to_section TEXT,
                            signatory_name TEXT,
                            signatory_designation TEXT
                        );
                    """)
                    cursor.execute("""
                        INSERT INTO special_events (id, emp_id, event_type, from_date, to_date, order_number, location, from_section, to_section, signatory_name, signatory_designation)
                        SELECT id, emp_id, event_type, from_date, to_date, order_number, location, from_section, to_section, signatory_name, signatory_designation FROM special_events_old;
                    """)
                
                cursor.execute(f"DROP TABLE {table_name}_old;")
                cursor.execute("PRAGMA foreign_keys = ON;")
                conn.commit()
                print(f"Table {table_name} repaired successfully.")
            except Exception as e:
                conn.rollback()
                print(f"Error repairing table {table_name}: {e}")

    # 1. Lines Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            line_name TEXT NOT NULL UNIQUE,
            color_code TEXT NOT NULL
        );
    """)
    
    # 2. Sections Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            line_id INTEGER REFERENCES lines(id) ON DELETE CASCADE,
            section_code TEXT NOT NULL UNIQUE,
            section_name TEXT NOT NULL,
            base_location TEXT NOT NULL
        );
    """)

    # 3. Employees Table
    cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='employees'")
    row = cursor.fetchone()
    if row and 'Flexible' not in row[0]:
        print("Migrating employees table to allow 'Flexible' as default_rest_day...")
        try:
            cursor.execute("PRAGMA foreign_keys = OFF;")
            cursor.execute("ALTER TABLE employees RENAME TO employees_old;")
            cursor.execute("""
                CREATE TABLE employees (
                    emp_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    pf_number TEXT NOT NULL UNIQUE,
                    name TEXT NOT NULL,
                    designation TEXT NOT NULL,
                    level INTEGER NOT NULL CHECK (level >= 1 AND level <= 12),
                    primary_section_id INTEGER REFERENCES sections(id) ON DELETE SET NULL,
                    default_rest_day TEXT NOT NULL CHECK (default_rest_day IN ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday', 'Flexible')),
                    joining_date TEXT,
                    weekly_schedule TEXT,
                    display_order INTEGER DEFAULT 0
                );
            """)
            cursor.execute("""
                INSERT INTO employees (emp_id, pf_number, name, designation, level, primary_section_id, default_rest_day, joining_date, weekly_schedule, display_order)
                SELECT emp_id, pf_number, name, designation, level, primary_section_id, default_rest_day, joining_date, weekly_schedule, display_order FROM employees_old;
            """)
            cursor.execute("DROP TABLE employees_old;")
            cursor.execute("PRAGMA foreign_keys = ON;")
            conn.commit()
            print("Migration of employees table successful.")
        except Exception as e:
            conn.rollback()
            print(f"Error migrating employees table: {e}")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            emp_id INTEGER PRIMARY KEY AUTOINCREMENT,
            pf_number TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            designation TEXT NOT NULL,
            level INTEGER NOT NULL CHECK (level >= 1 AND level <= 12),
            primary_section_id INTEGER REFERENCES sections(id) ON DELETE SET NULL,
            default_rest_day TEXT NOT NULL CHECK (default_rest_day IN ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday', 'Flexible')),
            joining_date TEXT,
            weekly_schedule TEXT,
            display_order INTEGER DEFAULT 0
        );
    """)

    # 4. Compensatory Rest Ledger Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS compensatory_rest_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id INTEGER REFERENCES employees(emp_id) ON DELETE CASCADE,
            earned_date TEXT NOT NULL,
            consumed_date TEXT,
            UNIQUE (emp_id, earned_date)
        );
    """)

    # 5. Shift Rules Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS shift_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            section_id INTEGER REFERENCES sections(id) ON DELETE CASCADE,
            shift_code TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            working_days TEXT NOT NULL,
            is_night_duty INTEGER NOT NULL DEFAULT 0,
            duty_type TEXT,
            UNIQUE (section_id, shift_code)
        );
    """)

    # 6. Leave Bank Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leave_bank (
            emp_id INTEGER REFERENCES employees(emp_id) ON DELETE CASCADE,
            year INTEGER NOT NULL,
            total_cl INTEGER NOT NULL DEFAULT 8,
            total_lap INTEGER NOT NULL DEFAULT 30,
            used_cl INTEGER NOT NULL DEFAULT 0,
            used_lap INTEGER NOT NULL DEFAULT 0,
            accrued_cr INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (emp_id, year)
        );
    """)

    # 7. Attendance Log Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id INTEGER REFERENCES employees(emp_id) ON DELETE CASCADE,
            date TEXT NOT NULL,
            status TEXT NOT NULL,
            is_night INTEGER NOT NULL DEFAULT 0,
            shift_id INTEGER REFERENCES shift_rules(id) ON DELETE SET NULL,
            remarks TEXT,
            UNIQUE (emp_id, date)
        );
    """)

    # 8. Special Events Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS special_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id INTEGER REFERENCES employees(emp_id) ON DELETE CASCADE,
            event_type TEXT NOT NULL,
            from_date TEXT NOT NULL,
            to_date TEXT NOT NULL,
            order_number TEXT,
            location TEXT,
            from_section TEXT,
            to_section TEXT,
            signatory_name TEXT,
            signatory_designation TEXT
        );
    """)

    # 9. Attendance Codes Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance_codes (
            code TEXT PRIMARY KEY,
            description TEXT NOT NULL,
            bg_color TEXT NOT NULL DEFAULT '#FFFFFF',
            text_color TEXT NOT NULL DEFAULT '#1E293B',
            is_leave INTEGER NOT NULL DEFAULT 0,
            leave_type TEXT NOT NULL CHECK (leave_type IN ('CL', 'LAP', 'CR', 'Sick', 'None')) DEFAULT 'None'
        );
    """)

    # 10. Holidays Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS holidays (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            holiday_date TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            holiday_type TEXT NOT NULL,
            applicability TEXT
        );
    """)

    # 11. Audit Logs Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            user TEXT NOT NULL,
            module TEXT NOT NULL,
            action TEXT NOT NULL,
            details TEXT NOT NULL
        );
    """)

    # Seed Default Roster Status Codes
    cursor.execute("SELECT count(*) FROM attendance_codes")
    if cursor.fetchone()[0] == 0:
        default_codes = [
            ('P', 'Present (General/Day Shift)', '#FFFFFF', '#1E293B', 0, 'None'),
            ('P/N', 'Present (Night Shift)', '#F3E8FF', '#7E22CE', 0, 'None'),
            ('R', 'Weekly Rest Day', '#F1F5F9', '#64748B', 0, 'None'),
            ('CR', 'Compensatory Rest', '#EFF6FF', '#1D4ED8', 1, 'CR'),
            ('CL', 'Casual Leave', '#FFFBEB', '#B45309', 1, 'CL'),
            ('LAP', 'Average Pay Leave (LAP)', '#FFF7ED', '#C2410C', 1, 'LAP'),
            ('Sick', 'Sick Leave / Medical Memo', '#FEF2F2', '#B91C1C', 1, 'Sick'),
            ('SCL', 'Special Casual Leave', '#FFF1F2', '#BE123C', 1, 'None'),
            ('PH', 'Public / National Holiday', '#FEF9C3', '#A16207', 0, 'None')
        ]
        cursor.executemany("""
            INSERT INTO attendance_codes (code, description, bg_color, text_color, is_leave, leave_type)
            VALUES (?, ?, ?, ?, ?, ?)
        """, default_codes)

    # Seed Default Lines and Sections
    cursor.execute("SELECT count(*) FROM lines")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO lines (line_name, color_code) VALUES ('Blue Line', '#005EA6')")
        cursor.execute("INSERT INTO lines (line_name, color_code) VALUES ('Yellow Line', '#FFD100')")
        cursor.execute("INSERT INTO lines (line_name, color_code) VALUES ('Green Line', '#009639')")
        cursor.execute("INSERT INTO lines (line_name, color_code) VALUES ('Purple Line', '#7B2E8D')")
        cursor.execute("INSERT INTO lines (line_name, color_code) VALUES ('Noapara Car Shed', '#475569')")

        cursor.execute("INSERT INTO sections (line_id, section_code, section_name, base_location) VALUES (1, 'KKVS', 'Kavi Subhash Section', 'Kavi Subhash')")
        cursor.execute("INSERT INTO sections (line_id, section_code, section_name, base_location) VALUES (1, 'KMUK', 'Tollygunge Section', 'Mahanayak Uttam Kumar')")
        cursor.execute("INSERT INTO sections (line_id, section_code, section_name, base_location) VALUES (1, 'KNAP', 'Noapara Section', 'Noapara')")
        cursor.execute("INSERT INTO sections (line_id, section_code, section_name, base_location) VALUES (2, 'KJHD', 'Joy Hind Section', 'Joy Hind')")
        cursor.execute("INSERT INTO sections (line_id, section_code, section_name, base_location) VALUES (4, 'KJKA', 'Joka Section', 'Joka')")
        cursor.execute("INSERT INTO sections (line_id, section_code, section_name, base_location) VALUES (5, 'KNCS', 'Noapara Car Shed Section', 'Noapara Car Shed')")
        
        # Seed default shift rules for all sections
        cursor.execute("SELECT id FROM sections")
        sec_ids = [r[0] for r in cursor.fetchall()]
        for sec_id in sec_ids:
            cursor.execute("INSERT INTO shift_rules (section_id, shift_code, start_time, end_time, working_days, is_night_duty) VALUES (?, 'M', '06:00:00', '14:00:00', 'Monday,Tuesday,Wednesday,Thursday,Friday,Saturday,Sunday', 0)", (sec_id,))
            cursor.execute("INSERT INTO shift_rules (section_id, shift_code, start_time, end_time, working_days, is_night_duty) VALUES (?, 'E', '14:00:00', '22:00:00', 'Monday,Tuesday,Wednesday,Thursday,Friday,Saturday,Sunday', 0)", (sec_id,))
            cursor.execute("INSERT INTO shift_rules (section_id, shift_code, start_time, end_time, working_days, is_night_duty) VALUES (?, 'N', '22:00:00', '06:00:00', 'Monday,Tuesday,Wednesday,Thursday,Friday,Saturday,Sunday', 1)", (sec_id,))
            cursor.execute("INSERT INTO shift_rules (section_id, shift_code, start_time, end_time, working_days, is_night_duty) VALUES (?, 'G', '09:00:00', '17:30:00', 'Monday,Tuesday,Wednesday,Thursday,Friday,Saturday', 0)", (sec_id,))

    conn.commit()

    # Auto-seed employees from local Excel list if employees list is currently empty
    cursor.execute("SELECT count(*) FROM employees")
    if cursor.fetchone()[0] == 0:
        excel_path = "Staff list for KKVS and KMUK section (1).xlsx"
        if not os.path.exists(excel_path):
            excel_path = os.path.join("..", excel_path)
        if os.path.exists(excel_path):
            try:
                df = pd.read_excel(excel_path, sheet_name="Staff List", header=None)
                current_section = "KKVS"
                for idx, row in df.iterrows():
                    if idx < 10:
                        continue
                    sl = row[0]
                    pf = row[1]
                    name = row[3]
                    desig = row[4]
                    if pd.isna(sl) and pd.isna(pf) and pd.isna(name) and pd.isna(desig):
                        continue
                    if str(sl).strip() in ["KKVS", "KMUK"]:
                        current_section = str(sl).strip()
                        continue
                    if pd.notna(pf) and pd.notna(name):
                        pf_str = str(pf).strip().split('.')[0]
                        name_str = str(name).replace("Sri ", "").replace("MD ", "Md. ").strip(" .")
                        desig_str = str(desig).strip(" .")
                        rest_day = "Wednesday" if current_section == "KKVS" else "Sunday"
                        
                        level = 1
                        if "sse" in desig_str.lower():
                            level = 8 if "ic" in desig_str.lower() else 7
                        elif "je" in desig_str.lower():
                            level = 6
                        elif "sr. tech" in desig_str.lower() or "sr.tech" in desig_str.lower():
                            level = 6
                        elif "tech -i" in desig_str.lower() or "tech-i" in desig_str.lower():
                            level = 5
                        elif "tech -ii" in desig_str.lower() or "tech-ii" in desig_str.lower():
                            level = 4
                        elif "tech -iii" in desig_str.lower() or "tech-iii" in desig_str.lower():
                            level = 3
                        
                        schedule = {
                            "Monday": "G", "Tuesday": "G", 
                            "Wednesday": "R" if rest_day == "Wednesday" else "G",
                            "Thursday": "G", "Friday": "G", "Saturday": "G",
                            "Sunday": "R" if rest_day == "Sunday" else "G"
                        }
                        
                        cursor.execute("""
                            INSERT INTO employees (pf_number, name, designation, level, primary_section_id, default_rest_day, weekly_schedule)
                            VALUES (?, ?, ?, ?, (SELECT id FROM sections WHERE section_code = ?), ?, ?)
                        """, (pf_str, name_str, desig_str, level, current_section, rest_day, json.dumps(schedule)))
                        
                        # Add initial leave bank
                        emp_id = cursor.lastrowid
                        cursor.execute("""
                            INSERT INTO leave_bank (emp_id, year, total_cl, total_lap, used_cl, used_lap, accrued_cr)
                            VALUES (?, 2026, 8, 30, 0, 0, 0)
                        """, (emp_id,))
                conn.commit()
                log_audit("Auto-Seed", "Database", "Initial database seeded with staff from Excel list.")
            except Exception as e:
                print("Excel seeding failed:", e)
                
    # Ensure display_order column exists in employees table (migration for existing DBs)
    try:
        cursor.execute("ALTER TABLE employees ADD COLUMN display_order INTEGER DEFAULT 0;")
        conn.commit()
    except sqlite3.OperationalError:
        pass # Column already exists

    # Ensure transfer and signatory columns exist in special_events table
    for col in ["from_section", "to_section", "signatory_name", "signatory_designation"]:
        try:
            cursor.execute(f"ALTER TABLE special_events ADD COLUMN {col} TEXT;")
            conn.commit()
        except sqlite3.OperationalError:
            pass # Column already exists
            
    # Ensure duty_type column exists in shift_rules table
    try:
        cursor.execute("ALTER TABLE shift_rules ADD COLUMN duty_type TEXT;")
        conn.commit()
    except sqlite3.OperationalError:
        pass # Column already exists

    # Ensure basic_pay column exists in employees table (migration for existing DBs)
    try:
        cursor.execute("ALTER TABLE employees ADD COLUMN basic_pay INTEGER DEFAULT 0;")
        conn.commit()
    except sqlite3.OperationalError:
        pass # Column already exists

    # Create ta_bills and ta_entries tables
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ta_bills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id INTEGER NOT NULL REFERENCES employees(emp_id) ON DELETE CASCADE,
            month_year TEXT NOT NULL,
            journey_type TEXT NOT NULL,
            book_no TEXT DEFAULT '',
            page_no TEXT DEFAULT '',
            serial_no_from TEXT DEFAULT '',
            serial_no_to TEXT DEFAULT '',
            bill_unit TEXT DEFAULT '',
            basic_pay INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        );
    """)
    try:
        cursor.execute("ALTER TABLE ta_bills ADD COLUMN bill_unit TEXT DEFAULT '';")
        conn.commit()
    except sqlite3.OperationalError:
        pass # Column already exists

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ta_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_id INTEGER NOT NULL REFERENCES ta_bills(id) ON DELETE CASCADE,
            entry_date TEXT NOT NULL,
            train_no TEXT DEFAULT '',
            time_left TEXT DEFAULT '',
            time_arrived TEXT DEFAULT '',
            station_from TEXT DEFAULT '',
            station_to TEXT DEFAULT '',
            is_stay INTEGER DEFAULT 0,
            stay_details TEXT DEFAULT '',
            days_nights TEXT DEFAULT '',
            object_journey TEXT DEFAULT '',
            rate INTEGER NOT NULL,
            amount INTEGER NOT NULL,
            display_order INTEGER DEFAULT 0
        );
    """)
    conn.commit()

    # Create roster_rules table and seed default KKVS rotation rule
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS roster_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            pattern TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
    """)
    cursor.execute("SELECT COUNT(*) FROM roster_rules")
    if cursor.fetchone()[0] == 0:
        cursor.execute("""
            INSERT INTO roster_rules (name, pattern)
            VALUES (?, ?)
        """, ("3-Week Rotating (KKVS)", "E,E,E,E,E,E,E,R,M,M,M,M,N,N,N,N,N,N,N,R,R"))
    conn.commit()
        
    conn.close()

init_db()

# --- Pydantic Models for API Requests ---
class LineSchema(BaseModel):
    line_name: str
    color_code: str

class SectionSchema(BaseModel):
    line_id: int
    section_code: str
    section_name: str
    base_location: str

class EmployeeSchema(BaseModel):
    pf_number: str
    name: str
    designation: str
    level: int
    primary_section_id: Optional[int] = None
    default_rest_day: str
    joining_date: Optional[str] = None
    weekly_schedule: Optional[dict] = None
    display_order: Optional[int] = 0
    basic_pay: Optional[int] = 0

class ReorderPayload(BaseModel):
    emp_ids: List[int]

class ShiftRuleSchema(BaseModel):
    section_id: int
    shift_code: str
    start_time: str
    end_time: str
    working_days: List[str]
    is_night_duty: bool
    duty_type: Optional[str] = None

class RosterRuleSchema(BaseModel):
    name: str
    pattern: str

class AttendanceCodeSchema(BaseModel):
    code: str
    description: str
    bg_color: str
    text_color: str
    is_leave: bool
    leave_type: str

class HolidaySchema(BaseModel):
    holiday_date: str
    name: str
    holiday_type: str
    applicability: Optional[str] = None

class AttendanceLogSchema(BaseModel):
    emp_id: int
    date: str
    status: str
    is_night: bool
    shift_id: Optional[int] = None
    remarks: Optional[str] = ""

class SpecialEventSchema(BaseModel):
    emp_id: int
    event_type: str
    from_date: str
    to_date: str
    order_number: Optional[str] = None
    location: Optional[str] = None
    from_section: Optional[str] = None
    to_section: Optional[str] = None
    signatory_name: Optional[str] = None
    signatory_designation: Optional[str] = None

class RestoreSchema(BaseModel):
    filename: str

class NightDutyRow(BaseModel):
    sl: int
    pf_number: str
    name: str
    designation: str
    level: int
    dates: str # Comma-separated days, e.g. "20,21,24,30"
    total_days: int
    section_code: Optional[str] = ""
    remarks: Optional[str] = ""

class NightDutyExportRequest(BaseModel):
    month_name: str # e.g. "APRIL-2026"
    section_code: str # KKVS, KMUK, etc.
    section_name: str # e.g. "Kavi Subhash Section"
    ref_no: str # e.g. "SSE/Sig/KKVS/05/26"
    bill_unit: str # e.g. "2201-806"
    date_str: str # e.g. "30.05.2026"
    signatory_left: str # "SSE/Sig/KKVS/IC"
    signatory_right: str # "Dy. CPO"
    rows: List[NightDutyRow]

class AttendanceDay(BaseModel):
    day: int
    weekday: str # "Mon", "Tue", etc.
    status: str # "P", "R", "CR", etc.
    is_holiday: bool = False
    remarks: Optional[str] = ""

class AttendanceRow(BaseModel):
    sl: int
    pf_number: str
    name: str
    designation: str
    days: List[AttendanceDay]
    section_code: Optional[str] = ""
    remarks: Optional[str] = ""

class AttendanceExportRequest(BaseModel):
    period_start: str # "11.05.2026"
    period_end: str # "10.06.2026"
    section_code: str # KKVS
    section_name: str # KAVI SUBHASH (KKVS)
    submission_date: str # 10.06.2026
    signatory_left: str # SSE/SIG/KKVS/IC
    signatory_right: str # Dy. CPO
    rows: List[AttendanceRow]

# --- Helper Functions for Weightage ---
def calculate_weightage(total_days: int):
    # 8 hours per shift. Weightage = 10 mins per hour = 80 mins per shift.
    total_mins = total_days * 80
    hours = total_mins // 60
    minutes = total_mins % 60
    return f"{hours:02d} HRS", f"{minutes:02d} MIN."

def calculate_weightage_numeric(total_days: int):
    total_mins = total_days * 80
    hours = total_mins // 60
    minutes = total_mins % 60
    return hours, minutes

# --- Numbered Canvas for PDF page counting ---
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super(NumberedCanvas, self).__init__(*args, **kwargs)
        self._saved_page_states = []
        self._pageCompression = 0

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#4A5568"))
        # Header
        self.drawString(36, self._pagesize[1] - 30, "Metro Railway, Kolkata — Signalling & Telecommunication Department")
        self.setLineWidth(0.5)
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.line(36, self._pagesize[1] - 35, self._pagesize[0] - 36, self._pagesize[1] - 35)
        
        # Footer
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(self._pagesize[0] - 36, 30, page_text)
        self.drawString(36, 30, f"Generated on {date.today().strftime('%d.%m.%Y')} | Official S&T ERP Report")
        self.line(36, 40, self._pagesize[0] - 36, 40)
        self.restoreState()


# --- API ENDPOINTS ---

try:
    from version import VERSION
except ImportError:
    VERSION = "1.2.2"

@app.get("/api/version")
def get_version():
    return {"version": VERSION}

# --- Background Updater State & Endpoints ---
import threading
import urllib.request
import json
import os

download_state = {
    "status": "idle", # "idle", "downloading", "completed", "error"
    "progress": 0,
    "filename": "",
    "path": "",
    "error_message": ""
}

def download_file_worker(download_url, dest_path):
    global download_state
    try:
        req = urllib.request.Request(
            download_url, 
            headers={"User-Agent": "StaffAttendanceERPUpdater"}
        )
        with urllib.request.urlopen(req) as response:
            total_size = int(response.headers.get('content-length', 0))
            block_size = 1024 * 16
            downloaded = 0
            
            with open(dest_path, "wb") as f:
                while True:
                    chunk = response.read(block_size)
                    if not chunk:
                        break
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total_size > 0:
                        download_state["progress"] = int((downloaded / total_size) * 100)
                    else:
                        download_state["progress"] = 50
            
            download_state["status"] = "completed"
            download_state["progress"] = 100
    except Exception as e:
        download_state["status"] = "error"
        download_state["error_message"] = str(e)
        print(f"Update download failed: {e}")

@app.post("/api/updater/download")
def trigger_update_download():
    global download_state
    if download_state["status"] == "downloading":
        return {"status": "already_downloading"}
        
    try:
        req = urllib.request.Request(
            "https://api.github.com/repos/kbiswas9876/StaffAttendanceNAD/releases/latest",
            headers={"User-Agent": "StaffAttendanceERPUpdater"}
        )
        with urllib.request.urlopen(req) as response:
            release_data = json.loads(response.read().decode())
            
        assets = release_data.get("assets", [])
        exe_asset = None
        for asset in assets:
            if asset.get("name", "").endswith(".exe"):
                exe_asset = asset
                break
                
        if not exe_asset:
            if assets:
                exe_asset = assets[0]
            else:
                raise HTTPException(status_code=404, detail="No assets found in the latest release.")
                
        download_url = exe_asset["browser_download_url"]
        filename = exe_asset["name"]
        
        downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        os.makedirs(downloads_dir, exist_ok=True)
        dest_path = os.path.join(downloads_dir, filename)
        
        download_state["status"] = "downloading"
        download_state["progress"] = 0
        download_state["filename"] = filename
        download_state["path"] = dest_path
        download_state["error_message"] = ""
        
        t = threading.Thread(target=download_file_worker, args=(download_url, dest_path), daemon=True)
        t.start()
        
        return {
            "status": "started",
            "filename": filename,
            "path": dest_path
        }
    except Exception as e:
        download_state["status"] = "error"
        download_state["error_message"] = str(e)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/updater/status")
def get_updater_status():
    global download_state
    return download_state


# 1. Lines CRUD
@app.get("/api/lines")
def read_lines():
    conn = get_db()
    lines = [dict(row) for row in conn.execute("SELECT * FROM lines").fetchall()]
    conn.close()
    return lines

@app.post("/api/lines")
def add_line(payload: LineSchema):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO lines (line_name, color_code) VALUES (?, ?)", (payload.line_name, payload.color_code))
        conn.commit()
        log_audit("Insert", "Lines", f"Created line: {payload.line_name}")
        return {"id": cursor.lastrowid, **payload.dict()}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Line name already exists.")
    finally:
        conn.close()

@app.put("/api/lines/{id}")
def edit_line(id: int, payload: LineSchema):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE lines SET line_name = ?, color_code = ? WHERE id = ?", (payload.line_name, payload.color_code, id))
    conn.commit()
    conn.close()
    log_audit("Update", "Lines", f"Updated line ID {id}: {payload.line_name}")
    return {"id": id, **payload.dict()}

@app.delete("/api/lines/{id}")
def remove_line(id: int):
    conn = get_db()
    conn.execute("DELETE FROM lines WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    log_audit("Delete", "Lines", f"Deleted line ID {id}")
    return {"status": "success"}

# 2. Sections CRUD
@app.get("/api/sections")
def read_sections():
    conn = get_db()
    sections = [dict(row) for row in conn.execute("SELECT * FROM sections").fetchall()]
    conn.close()
    return sections

@app.post("/api/sections")
def add_section(payload: SectionSchema):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO sections (line_id, section_code, section_name, base_location)
            VALUES (?, ?, ?, ?)
        """, (payload.line_id, payload.section_code, payload.section_name, payload.base_location))
        conn.commit()
        log_audit("Insert", "Sections", f"Created section: {payload.section_code}")
        return {"id": cursor.lastrowid, **payload.dict()}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Section code already exists.")
    finally:
        conn.close()

@app.put("/api/sections/{id}")
def edit_section(id: int, payload: SectionSchema):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE sections SET line_id = ?, section_code = ?, section_name = ?, base_location = ?
        WHERE id = ?
    """, (payload.line_id, payload.section_code, payload.section_name, payload.base_location, id))
    conn.commit()
    conn.close()
    log_audit("Update", "Sections", f"Updated section ID {id}: {payload.section_code}")
    return {"id": id, **payload.dict()}

@app.delete("/api/sections/{id}")
def remove_section(id: int):
    conn = get_db()
    conn.execute("DELETE FROM sections WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    log_audit("Delete", "Sections", f"Deleted section ID {id}")
    return {"status": "success"}

# 3. Employees CRUD
@app.get("/api/employees")
def read_employees(section_code: Optional[str] = None):
    conn = get_db()
    query = """
        SELECT e.*, s.section_code 
        FROM employees e
        LEFT JOIN sections s ON e.primary_section_id = s.id
    """
    if section_code:
        query += " WHERE s.section_code = ? ORDER BY e.display_order ASC, e.emp_id ASC"
        rows = conn.execute(query, (section_code,)).fetchall()
    else:
        query += " ORDER BY e.display_order ASC, e.emp_id ASC"
        rows = conn.execute(query).fetchall()
        
    employees = []
    for r in rows:
        d = dict(r)
        d['weekly_schedule'] = json.loads(d['weekly_schedule']) if d['weekly_schedule'] else {}
        employees.append(d)
    conn.close()
    return employees

@app.get("/api/employees/{emp_id}")
def read_employee(emp_id: int):
    conn = get_db()
    row = conn.execute("""
        SELECT e.*, s.section_code 
        FROM employees e
        LEFT JOIN sections s ON e.primary_section_id = s.id
        WHERE e.emp_id = ?
    """, (emp_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Employee not found")
    d = dict(row)
    d['weekly_schedule'] = json.loads(d['weekly_schedule']) if d['weekly_schedule'] else {}
    return d

@app.post("/api/employees")
def add_employee(payload: EmployeeSchema):
    conn = get_db()
    try:
        cursor = conn.cursor()
        sched_str = json.dumps(payload.weekly_schedule) if payload.weekly_schedule else "{}"
        cursor.execute("""
            INSERT INTO employees (pf_number, name, designation, level, primary_section_id, default_rest_day, joining_date, weekly_schedule, display_order, basic_pay)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (payload.pf_number, payload.name, payload.designation, payload.level, payload.primary_section_id, payload.default_rest_day, payload.joining_date, sched_str, payload.display_order or 0, payload.basic_pay or 0))
        emp_id = cursor.lastrowid
        
        # Create initial leave bank record
        cursor.execute("""
            INSERT INTO leave_bank (emp_id, year, total_cl, total_lap, used_cl, used_lap, accrued_cr)
            VALUES (?, 2026, 8, 30, 0, 0, 0)
        """, (emp_id,))
        conn.commit()
        log_audit("Insert", "Employees", f"Enrolled employee {payload.name} (PF: {payload.pf_number})")
        return {"emp_id": emp_id, **payload.dict()}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="PF Number already exists.")
    finally:
        conn.close()

@app.put("/api/employees/{emp_id}")
def edit_employee(emp_id: int, payload: EmployeeSchema):
    conn = get_db()
    cursor = conn.cursor()
    sched_str = json.dumps(payload.weekly_schedule) if payload.weekly_schedule else "{}"
    cursor.execute("""
        UPDATE employees SET pf_number = ?, name = ?, designation = ?, level = ?, 
            primary_section_id = ?, default_rest_day = ?, joining_date = ?, weekly_schedule = ?, display_order = ?, basic_pay = ?
        WHERE emp_id = ?
    """, (payload.pf_number, payload.name, payload.designation, payload.level, payload.primary_section_id, payload.default_rest_day, payload.joining_date, sched_str, payload.display_order or 0, payload.basic_pay or 0, emp_id))
    conn.commit()
    conn.close()
    
    # Recalculate leaves for 2026
    sync_leave_and_ledger(emp_id, 2026)
    log_audit("Update", "Employees", f"Updated employee ID {emp_id}: {payload.name}")
    return {"emp_id": emp_id, **payload.dict()}

@app.delete("/api/employees/{emp_id}")
def remove_employee(emp_id: int):
    conn = get_db()
    conn.execute("DELETE FROM employees WHERE emp_id = ?", (emp_id,))
    conn.commit()
    conn.close()
    log_audit("Delete", "Employees", f"Deleted employee ID {emp_id}")
    return {"status": "success"}

@app.post("/api/employees/reorder")
def reorder_employees(payload: ReorderPayload):
    conn = get_db()
    cursor = conn.cursor()
    try:
        for idx, emp_id in enumerate(payload.emp_ids):
            cursor.execute("UPDATE employees SET display_order = ? WHERE emp_id = ?", (idx, emp_id))
        conn.commit()
        log_audit("Reorder", "Employees", f"Reordered {len(payload.emp_ids)} employees")
        return {"status": "success"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# 4. Shift Rules
@app.get("/api/shift-rules")
def read_shift_rules(section_code: str):
    conn = get_db()
    rules = [dict(row) for row in conn.execute("""
        SELECT r.* FROM shift_rules r
        JOIN sections s ON r.section_id = s.id
        WHERE s.section_code = ?
    """, (section_code,)).fetchall()]
    for r in rules:
        r['working_days'] = r['working_days'].split(',')
        r['is_night_duty'] = bool(r['is_night_duty'])
        if not r.get('duty_type'):
            r['duty_type'] = 'Night Shift' if r['is_night_duty'] else 'General / Day Shift'
    conn.close()
    return rules

@app.post("/api/shift-rules")
def add_shift_rule(payload: ShiftRuleSchema):
    conn = get_db()
    try:
        cursor = conn.cursor()
        days_str = ",".join(payload.working_days)
        duty_type_val = payload.duty_type
        if not duty_type_val:
            duty_type_val = 'Night Shift' if payload.is_night_duty else 'General / Day Shift'
        cursor.execute("""
            INSERT INTO shift_rules (section_id, shift_code, start_time, end_time, working_days, is_night_duty, duty_type)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (payload.section_id, payload.shift_code, payload.start_time, payload.end_time, days_str, int(payload.is_night_duty), duty_type_val))
        conn.commit()
        log_audit("Insert", "Shift Rules", f"Added shift rule {payload.shift_code} for Section ID {payload.section_id}")
        return {"id": cursor.lastrowid, **payload.dict()}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Shift code already exists for this section.")
    finally:
        conn.close()

@app.put("/api/shift-rules/{rule_id}")
def update_shift_rule(rule_id: int, payload: ShiftRuleSchema):
    conn = get_db()
    try:
        cursor = conn.cursor()
        existing = cursor.execute("SELECT id FROM shift_rules WHERE id = ?", (rule_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Shift rule not found.")
        days_str = ",".join(payload.working_days)
        duty_type_val = payload.duty_type
        if not duty_type_val:
            duty_type_val = 'Night Shift' if payload.is_night_duty else 'General / Day Shift'
        cursor.execute("""
            UPDATE shift_rules 
            SET section_id = ?, shift_code = ?, start_time = ?, end_time = ?, working_days = ?, is_night_duty = ?, duty_type = ?
            WHERE id = ?
        """, (payload.section_id, payload.shift_code, payload.start_time, payload.end_time, days_str, int(payload.is_night_duty), duty_type_val, rule_id))
        conn.commit()
        log_audit("Update", "Shift Rules", f"Updated shift rule {payload.shift_code} (ID: {rule_id}) for Section ID {payload.section_id}")
        return {"id": rule_id, **payload.dict()}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Shift code already exists for this section.")
    finally:
        conn.close()

@app.delete("/api/shift-rules/{rule_id}")
def delete_shift_rule(rule_id: int):
    conn = get_db()
    try:
        cursor = conn.cursor()
        existing = cursor.execute("SELECT id, shift_code, section_id FROM shift_rules WHERE id = ?", (rule_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Shift rule not found.")
        cursor.execute("DELETE FROM shift_rules WHERE id = ?", (rule_id,))
        conn.commit()
        log_audit("Delete", "Shift Rules", f"Deleted shift rule {existing['shift_code']} (ID: {rule_id})")
        return {"status": "success"}
    finally:
        conn.close()

# 4.5. Roster Rules
@app.get("/api/roster-rules")
def read_roster_rules():
    conn = get_db()
    try:
        rules = [dict(row) for row in conn.execute("SELECT * FROM roster_rules ORDER BY name").fetchall()]
        return rules
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.post("/api/roster-rules")
def add_roster_rule(payload: RosterRuleSchema):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO roster_rules (name, pattern)
            VALUES (?, ?)
        """, (payload.name, payload.pattern))
        conn.commit()
        log_audit("Insert", "Roster Rules", f"Added roster rule {payload.name}")
        return {"id": cursor.lastrowid, **payload.dict()}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Roster rule name already exists.")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.delete("/api/roster-rules/{rule_id}")
def delete_roster_rule(rule_id: int):
    conn = get_db()
    try:
        conn.execute("DELETE FROM roster_rules WHERE id = ?", (rule_id,))
        conn.commit()
        log_audit("Delete", "Roster Rules", f"Deleted roster rule ID {rule_id}")
        return {"success": True}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# 5. Attendance Codes
@app.get("/api/attendance-codes")
def read_attendance_codes():
    conn = get_db()
    codes = [dict(row) for row in conn.execute("SELECT * FROM attendance_codes").fetchall()]
    for c in codes:
        c['is_leave'] = bool(c['is_leave'])
    conn.close()
    return codes

@app.post("/api/attendance-codes")
def add_attendance_code(payload: AttendanceCodeSchema):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO attendance_codes (code, description, bg_color, text_color, is_leave, leave_type)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (payload.code, payload.description, payload.bg_color, payload.text_color, int(payload.is_leave), payload.leave_type))
        conn.commit()
        log_audit("Insert", "Attendance Codes", f"Created roster code: {payload.code}")
        return payload.dict()
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Roster status code already exists.")
    finally:
        conn.close()

@app.put("/api/attendance-codes/{code}")
def edit_attendance_code(code: str, payload: AttendanceCodeSchema):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE attendance_codes SET description = ?, bg_color = ?, text_color = ?, is_leave = ?, leave_type = ?
        WHERE code = ?
    """, (payload.description, payload.bg_color, payload.text_color, int(payload.is_leave), payload.leave_type, code))
    conn.commit()
    conn.close()
    log_audit("Update", "Attendance Codes", f"Updated roster code: {code}")
    return payload.dict()

@app.delete("/api/attendance-codes/{code}")
def remove_attendance_code(code: str):
    conn = get_db()
    conn.execute("DELETE FROM attendance_codes WHERE code = ?", (code,))
    conn.commit()
    conn.close()
    log_audit("Delete", "Attendance Codes", f"Deleted roster code: {code}")
    return {"status": "success"}

# 6. CR Ledger
@app.get("/api/compensatory-rest-ledger/{emp_id}")
def read_cr_ledger(emp_id: int):
    conn = get_db()
    rows = [dict(row) for row in conn.execute("SELECT * FROM compensatory_rest_ledger WHERE emp_id = ?", (emp_id,)).fetchall()]
    conn.close()
    return rows

@app.post("/api/compensatory-rest-ledger")
def add_cr_ledger_entry(payload: dict):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO compensatory_rest_ledger (emp_id, earned_date, consumed_date)
            VALUES (?, ?, NULL)
        """, (payload['emp_id'], payload['earned_date']))
        conn.commit()
        log_audit("Insert", "CR Ledger", f"Manually logged extra-duty earned CR on {payload['earned_date']} for Employee ID {payload['emp_id']}")
        return {"id": cursor.lastrowid, **payload}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="CR record already exists for this date.")
    finally:
        conn.close()

@app.delete("/api/compensatory-rest-ledger/{id}")
def remove_cr_ledger_entry(id: int):
    conn = get_db()
    conn.execute("DELETE FROM compensatory_rest_ledger WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    log_audit("Delete", "CR Ledger", f"Deleted CR entry ID {id}")
    return {"status": "success"}

@app.put("/api/compensatory-rest-ledger/{id}")
def update_cr_ledger_consumed(id: int, payload: dict):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE compensatory_rest_ledger SET consumed_date = ? WHERE id = ?
    """, (payload.get('consumed_date'), id))
    conn.commit()
    conn.close()
    return {"status": "success"}

# 7. Leave Bank
@app.get("/api/leave-bank/{emp_id}")
def read_leave_bank(emp_id: int, year: int = 2026):
    conn = get_db()
    row = conn.execute("SELECT * FROM leave_bank WHERE emp_id = ? AND year = ?", (emp_id, year)).fetchone()
    if not row:
        # Create default
        cursor = conn.cursor()
        cursor.execute("INSERT INTO leave_bank (emp_id, year, total_cl, total_lap, used_cl, used_lap, accrued_cr) VALUES (?, ?, 8, 30, 0, 0, 0)", (emp_id, year))
        conn.commit()
        row = conn.execute("SELECT * FROM leave_bank WHERE emp_id = ? AND year = ?", (emp_id, year)).fetchone()
    conn.close()
    return dict(row)

@app.put("/api/leave-bank")
def edit_leave_bank(payload: dict):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO leave_bank (emp_id, year, total_cl, total_lap, used_cl, used_lap, accrued_cr)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(emp_id, year) DO UPDATE SET
            total_cl = excluded.total_cl,
            total_lap = excluded.total_lap,
            used_cl = excluded.used_cl,
            used_lap = excluded.used_lap,
            accrued_cr = excluded.accrued_cr
    """, (payload['emp_id'], payload['year'], payload['total_cl'], payload['total_lap'], payload['used_cl'], payload['used_lap'], payload['accrued_cr']))
    conn.commit()
    conn.close()
    log_audit("Update", "Leave Bank", f"Adjusted leave balances for Employee ID {payload['emp_id']}")
    return payload

# 8. Attendance Logs (Cycle Retrieval)
@app.get("/api/attendance-log")
def read_attendance_logs(section_code: str, start_date: str, end_date: str):
    conn = get_db()
    query = """
        SELECT l.* FROM attendance_log l
        JOIN employees e ON l.emp_id = e.emp_id
        JOIN sections s ON e.primary_section_id = s.id
        WHERE s.section_code = ? AND l.date >= ? AND l.date <= ?
    """
    rows = [dict(row) for row in conn.execute(query, (section_code, start_date, end_date)).fetchall()]
    for r in rows:
        r['is_night'] = bool(r['is_night'])
    conn.close()
    return rows

@app.get("/api/attendance-log/{emp_id}")
def read_employee_attendance_logs(emp_id: int, year: int = 2026):
    conn = get_db()
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    query = """
        SELECT * FROM attendance_log
        WHERE emp_id = ? AND date >= ? AND date <= ?
        ORDER BY date ASC
    """
    rows = [dict(row) for row in conn.execute(query, (emp_id, start_date, end_date)).fetchall()]
    for r in rows:
        r['is_night'] = bool(r['is_night'])
    conn.close()
    return rows


@app.post("/api/attendance-log")
def add_attendance_log(payload: AttendanceLogSchema):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO attendance_log (emp_id, date, status, is_night, shift_id, remarks)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(emp_id, date) DO UPDATE SET
            status = excluded.status,
            is_night = excluded.is_night,
            shift_id = excluded.shift_id,
            remarks = excluded.remarks
    """, (payload.emp_id, payload.date, payload.status, int(payload.is_night), payload.shift_id, payload.remarks))
    conn.commit()
    conn.close()
    
    # Recalculate leaves
    year = int(payload.date.split("-")[0])
    sync_leave_and_ledger(payload.emp_id, year)
    log_audit("Update", "Attendance", f"Logged attendance for Employee ID {payload.emp_id} on {payload.date} as {payload.status}")
    return {"status": "success"}

@app.post("/api/attendance-log/bulk")
def add_attendance_logs_bulk(logs: List[AttendanceLogSchema]):
    conn = get_db()
    cursor = conn.cursor()
    emp_years = set()
    for log in logs:
        cursor.execute("""
            INSERT INTO attendance_log (emp_id, date, status, is_night, shift_id, remarks)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(emp_id, date) DO UPDATE SET
                status = excluded.status,
                is_night = excluded.is_night,
                shift_id = excluded.shift_id,
                remarks = excluded.remarks
        """, (log.emp_id, log.date, log.status, int(log.is_night), log.shift_id, log.remarks))
        
        year = int(log.date.split("-")[0])
        emp_years.add((log.emp_id, year))
        
    conn.commit()
    conn.close()
    
    # Recalculate logs triggers
    for emp_id, year in emp_years:
        sync_leave_and_ledger(emp_id, year)
        
    log_audit("Update", "Attendance", f"Bulk logged {len(logs)} attendance roster records.")
    return {"status": "success"}

@app.delete("/api/attendance-log")
def delete_attendance_log(emp_id: int = Query(...), date: str = Query(...)):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT count(*) FROM attendance_log WHERE emp_id = ? AND date = ?", (emp_id, date))
        if cursor.fetchone()[0] == 0:
            raise HTTPException(status_code=404, detail="Log entry not found")
        cursor.execute("DELETE FROM attendance_log WHERE emp_id = ? AND date = ?", (emp_id, date))
        conn.commit()
        
        # Recalculate leaves
        year = int(date.split("-")[0])
        sync_leave_and_ledger(emp_id, year)
        log_audit("Delete", "Attendance", f"Deleted attendance log for Employee ID {emp_id} on {date}")
        return {"status": "success"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.delete("/api/attendance-log/range")
def delete_attendance_logs_range(
    start_date: str = Query(...),
    end_date: str = Query(...),
    section_code: Optional[str] = Query(None)
):
    conn = get_db()
    try:
        cursor = conn.cursor()
        if section_code and section_code != 'ALL':
            cursor.execute("""
                SELECT emp_id FROM employees
                WHERE primary_section_id = (SELECT id FROM sections WHERE section_code = ?)
            """, (section_code,))
            affected_emps = [row['emp_id'] for row in cursor.fetchall()]
            if not affected_emps:
                return {"status": "success", "count": 0}
                
            placeholders = ",".join("?" for _ in affected_emps)
            query = f"""
                DELETE FROM attendance_log
                WHERE emp_id IN ({placeholders}) AND date >= ? AND date <= ?
            """
            cursor.execute(query, affected_emps + [start_date, end_date])
        else:
            cursor.execute("SELECT DISTINCT emp_id FROM attendance_log WHERE date >= ? AND date <= ?", (start_date, end_date))
            affected_emps = [row['emp_id'] for row in cursor.fetchall()]
            cursor.execute("DELETE FROM attendance_log WHERE date >= ? AND date <= ?", (start_date, end_date))
        
        conn.commit()
        
        # Recalculate leaves for all affected employees
        year = int(start_date.split("-")[0])
        for emp_id in affected_emps:
            sync_leave_and_ledger(emp_id, year)
            
        log_audit("Delete Range", "Attendance", f"Deleted attendance logs from {start_date} to {end_date} for section {section_code or 'ALL'}")
        return {"status": "success", "count": len(affected_emps)}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# 9. Special Events (Transfers & Training)
@app.get("/api/special-events")
def read_special_events(section_code: Optional[str] = None):
    conn = get_db()
    query = """
        SELECT ev.* FROM special_events ev
        JOIN employees e ON ev.emp_id = e.emp_id
        LEFT JOIN sections s ON e.primary_section_id = s.id
    """
    if section_code and section_code != 'ALL':
        query += " WHERE s.section_code = ?"
        rows = [dict(row) for row in conn.execute(query, (section_code,)).fetchall()]
    else:
        rows = [dict(row) for row in conn.execute(query).fetchall()]
    conn.close()
    return rows

@app.post("/api/special-events")
def add_special_event(payload: SpecialEventSchema):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO special_events (emp_id, event_type, from_date, to_date, order_number, location, from_section, to_section, signatory_name, signatory_designation)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (payload.emp_id, payload.event_type, payload.from_date, payload.to_date, payload.order_number, payload.location, payload.from_section, payload.to_section, payload.signatory_name, payload.signatory_designation))
    conn.commit()
    conn.close()
    log_audit("Insert", "Special Events", f"Logged event {payload.event_type} for Employee ID {payload.emp_id}")
    return {"id": cursor.lastrowid, **payload.dict()}

@app.delete("/api/special-events/{id}")
def remove_special_event(id: int):
    conn = get_db()
    conn.execute("DELETE FROM special_events WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    log_audit("Delete", "Special Events", f"Deleted event record ID {id}")
    return {"status": "success"}

# 10. Holiday Master
@app.get("/api/holidays")
def read_holidays():
    conn = get_db()
    rows = [dict(row) for row in conn.execute("SELECT * FROM holidays ORDER BY holiday_date ASC").fetchall()]
    conn.close()
    return rows

@app.post("/api/holidays")
def add_holiday(payload: HolidaySchema):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO holidays (holiday_date, name, holiday_type, applicability)
            VALUES (?, ?, ?, ?)
        """, (payload.holiday_date, payload.name, payload.holiday_type, payload.applicability))
        conn.commit()
        log_audit("Insert", "Holidays", f"Created holiday: {payload.name} ({payload.holiday_date})")
        return {"id": cursor.lastrowid, **payload.dict()}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Holiday date already registered.")
    finally:
        conn.close()

@app.put("/api/holidays/{id}")
def edit_holiday(id: int, payload: HolidaySchema):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE holidays SET holiday_date = ?, name = ?, holiday_type = ?, applicability = ?
        WHERE id = ?
    """, (payload.holiday_date, payload.name, payload.holiday_type, payload.applicability, id))
    conn.commit()
    conn.close()
    log_audit("Update", "Holidays", f"Updated holiday ID {id}: {payload.name}")
    return {"id": id, **payload.dict()}

@app.delete("/api/holidays/{id}")
def remove_holiday(id: int):
    conn = get_db()
    conn.execute("DELETE FROM holidays WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    log_audit("Delete", "Holidays", f"Deleted holiday record ID {id}")
    return {"status": "success"}

# 11. Audit Logs Endpoint
@app.get("/api/audit-logs")
def read_audit_logs():
    conn = get_db()
    logs = [dict(row) for row in conn.execute("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 500").fetchall()]
    conn.close()
    return logs

# 12. Database Backup and Restore
@app.get("/api/backups")
def list_backups():
    files = [f for f in os.listdir(BACKUP_DIR) if f.endswith(".db")]
    files.sort(reverse=True)
    return files

@app.post("/api/backups/create")
def create_backup():
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"backup_{timestamp}.db"
        dest = os.path.join(BACKUP_DIR, backup_name)
        
        # Verify db integrity check before backing up
        conn = get_db()
        check = conn.execute("PRAGMA integrity_check;").fetchone()[0]
        conn.close()
        
        if check != "ok":
            raise HTTPException(status_code=500, detail="Database integrity check failed. Cannot backup corrupt file.")
            
        shutil.copy2(DB_PATH, dest)
        log_audit("Backup", "System", f"Created database backup snapshot: {backup_name}")
        return {"status": "success", "filename": backup_name}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Backup failed: {str(e)}")

@app.post("/api/backups/restore")
def restore_backup(payload: RestoreSchema):
    try:
        src = os.path.join(BACKUP_DIR, payload.filename)
        if not os.path.exists(src):
            raise HTTPException(status_code=404, detail="Backup file not found.")
            
        # Verify backup file integrity
        conn = sqlite3.connect(src)
        check = conn.execute("PRAGMA integrity_check;").fetchone()[0]
        conn.close()
        
        if check != "ok":
            raise HTTPException(status_code=500, detail="Backup file integrity verification failed. Restore aborted.")
            
        # Make a safety backup of current state
        if os.path.exists(DB_PATH):
            shutil.copy2(DB_PATH, os.path.join(BACKUP_DIR, "pre_restore_safety.db"))
        
        # Replace
        shutil.copy2(src, DB_PATH)
        log_audit("Restore", "System", f"Restored database from snapshot: {payload.filename}")
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")

@app.delete("/api/backups/{filename}")
def delete_backup(filename: str):
    try:
        filename = os.path.basename(filename)
        file_path = os.path.join(BACKUP_DIR, filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Backup file not found.")
        os.remove(file_path)
        log_audit("Delete Backup", "System", f"Permanently deleted database backup: {filename}")
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Delete backup failed: {str(e)}")

@app.get("/api/backups/status")
def get_backup_status():
    conn = get_db()
    integrity = conn.execute("PRAGMA integrity_check;").fetchone()[0]
    conn.close()
    
    last_backup = "None"
    files = [f for f in os.listdir(BACKUP_DIR) if f.endswith(".db") and f.startswith("backup_")]
    if files:
        files.sort(reverse=True)
        last_backup = files[0]
        
    return {
        "integrity": integrity,
        "last_backup": last_backup,
        "database_size_bytes": os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    }


# ==========================================
# --- EXCEL/PDF EXPORT SERVICES ---
# ==========================================

# --- Endpoint 1: Export Night Duty Excel ---
@app.post("/api/export/night-duty/excel")
async def export_night_duty_excel(req: NightDutyExportRequest):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    sheet = workbook.add_worksheet("Night Duty Statement")

    # --- Page Setup ---
    sheet.set_landscape()
    sheet.set_paper(9)  # A4
    sheet.set_margins(0.4, 0.4, 0.5, 0.5)
    sheet.set_header('&C&"Segoe UI,Bold"&10METRO RAILWAY, KOLKATA — SIGNALLING & TELECOMMUNICATION DEPT.')
    sheet.set_footer('&L&8Generated on ' + date.today().strftime('%d.%m.%Y') + '&R&8Page &P of &N')

    # --- Formats ---
    org_title_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 14, 'bold': True,
        'align': 'center', 'valign': 'vcenter', 'font_color': '#1E3A8A'
    })
    org_sub_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 10, 'bold': False,
        'align': 'center', 'valign': 'vcenter', 'font_color': '#374151'
    })
    letter_key_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter'
    })
    letter_val_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 10, 'align': 'left', 'valign': 'vcenter'
    })
    sub_banner_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 11, 'bold': True,
        'align': 'center', 'valign': 'vcenter',
        'bg_color': '#EFF6FF', 'font_color': '#1E3A8A',
        'border': 1, 'border_color': '#BFDBFE'
    })
    header_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 9, 'bold': True,
        'border': 1, 'border_color': '#CBD5E1',
        'bg_color': '#1E3A8A', 'font_color': '#FFFFFF',
        'align': 'center', 'valign': 'vcenter', 'text_wrap': True
    })
    data_center_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 9,
        'border': 1, 'border_color': '#E2E8F0',
        'align': 'center', 'valign': 'vcenter', 'text_wrap': True
    })
    data_alt_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 9,
        'border': 1, 'border_color': '#E2E8F0',
        'bg_color': '#F8FAFC',
        'align': 'center', 'valign': 'vcenter', 'text_wrap': True
    })
    name_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 9, 'bold': True,
        'border': 1, 'border_color': '#E2E8F0',
        'align': 'left', 'valign': 'vcenter'
    })
    name_alt_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 9, 'bold': True,
        'border': 1, 'border_color': '#E2E8F0',
        'bg_color': '#F8FAFC',
        'align': 'left', 'valign': 'vcenter'
    })
    wt_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 9, 'bold': True,
        'border': 1, 'border_color': '#E2E8F0',
        'bg_color': '#EFF6FF', 'font_color': '#1D4ED8',
        'align': 'center', 'valign': 'vcenter'
    })
    nil_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 9, 'italic': True,
        'border': 1, 'border_color': '#E2E8F0',
        'font_color': '#94A3B8', 'align': 'center', 'valign': 'vcenter'
    })
    sig_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 10, 'bold': True,
        'align': 'center', 'valign': 'vcenter',
        'top': 2, 'top_color': '#1E3A8A',
        'text_wrap': True
    })
    total_label_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 9, 'bold': True,
        'border': 1, 'border_color': '#CBD5E1',
        'bg_color': '#F1F5F9', 'font_color': '#1E3A8A',
        'align': 'right', 'valign': 'vcenter'
    })
    total_val_fmt = workbook.add_format({
        'font_name': 'Segoe UI', 'font_size': 9, 'bold': True,
        'border': 1, 'border_color': '#CBD5E1',
        'bg_color': '#F1F5F9', 'font_color': '#1E3A8A',
        'align': 'center', 'valign': 'vcenter'
    })

    # --- Column Widths ---
    # Cols: SL | PF | Name | Desig | Level | Dates | Total Days | Total Hours | Weightage Hrs | Weightage Mins | Remarks
    widths = [4, 16, 28, 14, 7, 60, 10, 12, 12, 12, 16]
    for col_idx, w in enumerate(widths):
        sheet.set_column(col_idx, col_idx, w)

    # ====== SECTION A: LETTERHEAD ======
    sheet.set_row(0, 30)
    sheet.merge_range(0, 0, 0, 10, "METRO RAILWAY, KOLKATA", org_title_fmt)
    sheet.set_row(1, 18)
    sheet.merge_range(1, 0, 1, 10, "Office of the Senior Divisional Signal & Telecommunication Engineer", org_sub_fmt)
    # Separator row
    sheet.set_row(2, 6)

    # ====== SECTION B: LETTER HEADER ======
    sheet.set_row(3, 18)
    sheet.write(3, 0, "No:", letter_key_fmt)
    sheet.merge_range(3, 1, 3, 4, req.ref_no, letter_val_fmt)
    sheet.write(3, 8, "Date:", letter_key_fmt)
    sheet.merge_range(3, 9, 3, 10, req.date_str, letter_val_fmt)

    sheet.set_row(4, 15)
    sheet.write(4, 0, "To,", letter_val_fmt)
    sheet.set_row(5, 15)
    sheet.merge_range(5, 0, 5, 10, f"The {req.signatory_right}, Metro Railway, Kolkata", letter_val_fmt)

    # ====== SECTION C: SUBJECT LINE ======
    sheet.set_row(6, 8)  # Small gap
    sheet.set_row(7, 30)
    sheet.merge_range(7, 0, 7, 10,
        f"Sub:  Night Duty Allowance Statement of Signal Staff (S&T Dept.) for the Month of {req.month_name}",
        sub_banner_fmt)
    sheet.set_row(8, 20)
    sheet.merge_range(8, 0, 8, 10,
        f"Ref:  Bill Unit No. {req.bill_unit}  |  Section: {req.section_name}",
        sub_banner_fmt)

    # ====== SECTION D: TABLE HEADERS ======
    sheet.set_row(9, 8)  # Gap row
    sheet.set_row(10, 20)
    sheet.set_row(11, 20)
    
    sheet.merge_range(10, 0, 11, 0, "SL\nNo.", header_fmt)
    sheet.merge_range(10, 1, 11, 1, "P.F. No.", header_fmt)
    sheet.merge_range(10, 2, 11, 2, "Name of Staff", header_fmt)
    sheet.merge_range(10, 3, 11, 3, "Designation", header_fmt)
    sheet.merge_range(10, 4, 11, 4, "Pay\nLevel", header_fmt)
    sheet.merge_range(10, 5, 11, 5, f"Dates of Night Duty\n({req.month_name})", header_fmt)
    sheet.merge_range(10, 6, 11, 6, "Total\nDays", header_fmt)
    sheet.merge_range(10, 7, 11, 7, "Total\nHours", header_fmt)
    
    # Horizontally merged header for Weightage
    sheet.merge_range(10, 8, 10, 9, "Weightage Time", header_fmt)
    sheet.write(11, 8, "Hrs", header_fmt)
    sheet.write(11, 9, "Mins", header_fmt)
    
    sheet.merge_range(10, 10, 11, 10, "Remarks", header_fmt)

    # ====== SECTION E: TABLE DATA ======
    nd_rows_with_data = 0
    current_row = 12
    
    is_joint_view = req.section_code == 'ALL'
    from collections import defaultdict
    grouped_rows = defaultdict(list)
    for row in req.rows:
        grouped_rows[row.section_code or 'KKVS'].append(row)
        
    for sec_code in sorted(grouped_rows.keys()):
        rows_in_sec = grouped_rows[sec_code]
        
        if is_joint_view:
            # Merged section header row
            section_title = f" SECTION: {sec_code.upper()} ({get_section_name_from_db(sec_code)})"
            sheet.merge_range(current_row, 0, current_row, 10, section_title, sub_banner_fmt)
            sheet.set_row(current_row, 22)
            current_row += 1
            
        for idx, row in enumerate(rows_in_sec):
            is_alt = idx % 2 == 1
            d_fmt = data_alt_fmt if is_alt else data_center_fmt
            n_fmt = name_alt_fmt if is_alt else name_fmt
            
            wt_hrs_val, wt_mins_val = calculate_weightage_numeric(row.total_days)
            total_hrs = row.total_days * 8
            has_dates = row.dates and row.dates.strip()
            
            # Estimate number of lines needed for the Dates string
            dates_str = row.dates or ""
            lines_needed = max(1, (len(dates_str) + 75) // 80)
            row_height = max(22, lines_needed * 14 + 6)
            
            sheet.set_row(current_row, row_height)
            sheet.write(current_row, 0, row.sl, d_fmt)
            sheet.write(current_row, 1, row.pf_number, d_fmt)
            sheet.write(current_row, 2, row.name, n_fmt)
            sheet.write(current_row, 3, row.designation, d_fmt)
            sheet.write(current_row, 4, row.level, d_fmt)
            
            # Dates with Nil for empty
            if has_dates:
                sheet.write(current_row, 5, row.dates, d_fmt)
                sheet.write(current_row, 6, row.total_days, d_fmt)
                sheet.write(current_row, 7, total_hrs, d_fmt)
                sheet.write(current_row, 8, wt_hrs_val, wt_fmt)
                sheet.write(current_row, 9, wt_mins_val, wt_fmt)
                nd_rows_with_data += 1
            else:
                sheet.write(current_row, 5, "Nil", nil_fmt)
                sheet.write(current_row, 6, 0, d_fmt)
                sheet.write(current_row, 7, 0, d_fmt)
                sheet.write(current_row, 8, "—", d_fmt)
                sheet.write(current_row, 9, "—", d_fmt)
            
            sheet.write(current_row, 10, row.remarks or "", d_fmt)
            current_row += 1

    # ====== SECTION F: TOTALS ROW ======
    sheet.set_row(current_row, 22)
    total_days_sum = sum(r.total_days for r in req.rows)
    total_hrs_sum = total_days_sum * 8
    sheet.merge_range(current_row, 0, current_row, 5, "TOTAL", total_label_fmt)
    sheet.write(current_row, 6, total_days_sum, total_val_fmt)
    sheet.write(current_row, 7, total_hrs_sum, total_val_fmt)
    wt_total_hrs_val, wt_total_mins_val = calculate_weightage_numeric(total_days_sum)
    sheet.write(current_row, 8, wt_total_hrs_val, total_val_fmt)
    sheet.write(current_row, 9, wt_total_mins_val, total_val_fmt)
    sheet.write(current_row, 10, f"{nd_rows_with_data} staff on night duty", total_label_fmt)
    current_row += 1

    # ====== SECTION G: SIGNATURES ======
    sheet.set_row(current_row, 8)
    current_row += 1
    sheet.set_row(current_row, 8)
    current_row += 1
    sheet.set_row(current_row, 8)
    sig_row = current_row + 3
    sheet.set_row(sig_row, 45)
    sheet.merge_range(sig_row, 0, sig_row, 3, req.signatory_left, sig_fmt)
    sheet.merge_range(sig_row, 7, sig_row, 10, f"For {req.signatory_right}", sig_fmt)
    sheet.set_row(sig_row + 1, 16)
    sheet.merge_range(sig_row + 1, 0, sig_row + 1, 3, "S&T Dept., Metro Railway, Kolkata",
                      workbook.add_format({'font_name': 'Segoe UI', 'font_size': 9, 'align': 'center', 'font_color': '#64748B'}))
    sheet.merge_range(sig_row + 1, 7, sig_row + 1, 10, "Metro Railway, Kolkata",
                      workbook.add_format({'font_name': 'Segoe UI', 'font_size': 9, 'align': 'center', 'font_color': '#64748B'}))

    workbook.close()
    output.seek(0)
    
    filename = f"Night_Duty_{req.section_code}_{req.month_name}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )

# --- Endpoint 2: Export Night Duty PDF ---
@app.post("/api/export/night-duty/pdf")
async def export_night_duty_pdf(req: NightDutyExportRequest):
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        rightMargin=36,
        leftMargin=36,
        topMargin=54,
        bottomMargin=54
    )

    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=14,
        alignment=1, # Center
        textColor=colors.HexColor("#1B365D")
    )
    normal_style = ParagraphStyle(
        'DocNormal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11
    )
    bold_style = ParagraphStyle(
        'DocBold',
        parent=normal_style,
        fontName='Helvetica-Bold'
    )
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        alignment=1, # Center
        textColor=colors.white
    )

    story = []

    # Title Banner
    story.append(Paragraph("METRO RAILWAY, KOLKATA", title_style))
    story.append(Paragraph("SIGNALLING & TELECOMMUNICATION DEPARTMENT", ParagraphStyle('DocSubTitle', parent=title_style, fontSize=10)))
    story.append(Spacer(1, 15))

    # Details Block
    details_data = [
        [Paragraph(f"<b>No:</b> {req.ref_no}", normal_style), Paragraph(f"<b>Date:</b> {req.date_str}", normal_style)],
        [Paragraph("<b>To,</b>", normal_style), ""],
        [Paragraph(f"The {req.signatory_right},", normal_style), ""],
        [Paragraph("Metro Railway, Kolkata", normal_style), ""]
    ]
    details_table = Table(details_data, colWidths=[400, 350])
    details_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('SPAN', (0,1), (1,1)),
        ('SPAN', (0,2), (1,2)),
        ('SPAN', (0,3), (1,3)),
    ]))
    story.append(details_table)
    story.append(Spacer(1, 15))

    # Subject line
    story.append(Paragraph(f"<b>Sub:</b> Night Duty Statement of Signal Staffs of S&T Deptt. for the Month of <b>{req.month_name}</b>", normal_style))
    story.append(Paragraph(f"<b>Ref:</b> Bill Unit No. {req.bill_unit}", normal_style))
    story.append(Spacer(1, 10))

    # Table Grid
    table_headers = [
        Paragraph("SL", header_style),
        Paragraph("P.F. No.", header_style),
        Paragraph("Name of Staff", header_style),
        Paragraph("Desig", header_style),
        Paragraph("Level", header_style),
        Paragraph(f"Date of Night Duty in {req.month_name}", header_style),
        Paragraph("Total Days", header_style),
        Paragraph("Total Hours", header_style),
        Paragraph("Weightage Hrs", header_style),
        Paragraph("Weightage Mins", header_style),
        Paragraph("Remarks", header_style)
    ]
    
    is_joint_view = req.section_code == 'ALL'
    from collections import defaultdict
    grouped_rows = defaultdict(list)
    for row in req.rows:
        grouped_rows[row.section_code or 'KKVS'].append(row)
        
    table_data = [table_headers]
    t_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1B365D")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (2,1), (2,-1), 'LEFT'), # Left align names
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
    ]
    
    section_title_style = ParagraphStyle(
        'PDFSectionTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=10.5,
        textColor=colors.HexColor("#1E293B")
    )
    
    current_r_idx = 1
    for sec_code in sorted(grouped_rows.keys()):
        rows_in_sec = grouped_rows[sec_code]
        
        if is_joint_view:
            banner_text = f"<b>SECTION: {sec_code.upper()}</b> ({get_section_name_from_db(sec_code)})"
            table_data.append([
                Paragraph(banner_text, section_title_style),
                "", "", "", "", "", "", "", "", "", ""
            ])
            t_style.append(('SPAN', (0, current_r_idx), (-1, current_r_idx)))
            t_style.append(('BACKGROUND', (0, current_r_idx), (-1, current_r_idx), colors.HexColor("#E2E8F0")))
            t_style.append(('ALIGN', (0, current_r_idx), (-1, current_r_idx), 'LEFT'))
            current_r_idx += 1
            
        for idx, row in enumerate(rows_in_sec):
            wt_hrs, wt_mins = calculate_weightage(row.total_days)
            tot_hrs = row.total_days * 8
            has_dates = row.dates and row.dates.strip()
            table_data.append([
                Paragraph(str(row.sl), normal_style),
                Paragraph(row.pf_number, normal_style),
                Paragraph(row.name, normal_style),
                Paragraph(row.designation, normal_style),
                Paragraph(str(row.level), normal_style),
                Paragraph(row.dates or "Nil", normal_style),
                Paragraph(str(row.total_days), normal_style),
                Paragraph(f"{tot_hrs} Hrs", normal_style),
                Paragraph(str(wt_hrs) if has_dates else "—", normal_style),
                Paragraph(str(wt_mins) if has_dates else "—", normal_style),
                Paragraph(row.remarks or "", normal_style)
            ])
            row_bg = colors.HexColor("#F8FAFC") if idx % 2 == 1 else colors.white
            t_style.append(('BACKGROUND', (0, current_r_idx), (-1, current_r_idx), row_bg))
            current_r_idx += 1

    grid_table = Table(table_data, colWidths=[20, 75, 105, 60, 35, 210, 40, 45, 45, 45, 90])
    grid_table.setStyle(TableStyle(t_style))
    story.append(grid_table)
    story.append(Spacer(1, 30))

    # Signature blocks
    sig_data = [
        [Paragraph(f"<b>{req.signatory_left}</b>", normal_style), Paragraph(f"<b>For {req.signatory_right}</b>", ParagraphStyle('RightSig', parent=normal_style, alignment=2))],
        ["S&T Dept., Metro Railway, Kolkata", "Metro Railway, Kolkata"]
    ]
    sig_table = Table(sig_data, colWidths=[400, 370])
    story.append(sig_table)

    doc.build(story, canvasmaker=NumberedCanvas)
    pdf_buffer.seek(0)
    
    filename = f"Night_Duty_{req.section_code}_{req.month_name}.pdf"
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )

# --- Endpoint 3: Export Attendance Sheet Excel ---
@app.post("/api/export/attendance/excel")
async def export_attendance_excel(req: AttendanceExportRequest):
    output   = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    sheet    = workbook.add_worksheet("Attendance")

    # ── Page & print setup ───────────────────────────────────────────────────
    sheet.set_landscape()
    sheet.set_paper(9)                  # A4
    sheet.set_margins(0.4, 0.4, 0.5, 0.5)
    sheet.fit_to_pages(1, 0)            # fit width to 1 page

    # ── Dynamic column layout ─────────────────────────────────────────────────
    # Compute how many day columns we actually need, then place Remarks
    # immediately after — no empty gap column ever.
    days_in_grid = req.rows[0].days if req.rows else []
    NUM_DAYS     = len(days_in_grid)    # typically 28–31
    RC           = 3 + NUM_DAYS         # Remarks column index (0-based)
    LC           = RC                   # alias — last column index

    # ── Column widths ─────────────────────────────────────────────────────────
    sheet.set_column(0, 0,  4.5)        # SL
    sheet.set_column(1, 1, 24.0)        # Name / PF
    sheet.set_column(2, 2, 14.0)        # Designation
    if NUM_DAYS > 0:
        sheet.set_column(3, RC - 1, 5.8)   # Day columns — slightly wider for readability
    sheet.set_column(RC, RC, 26.0)      # Remarks — immediately after last day

    # ── Row heights ───────────────────────────────────────────────────────────
    sheet.set_row(0, 28)   # Title
    sheet.set_row(1, 22)   # Subtitle
    sheet.set_row(2, 22)   # Date numbers  — taller for readability
    sheet.set_row(3, 18)   # Weekday abbrevs — taller

    # ── Font ──────────────────────────────────────────────────────────────────
    # Segoe UI — clean, modern, and highly legible system font
    FONT = 'Segoe UI'

    # ── Color palette ─────────────────────────────────────────────────────────
    # HEADER CHROME — Slate Charcoal Premium Theme
    C_H1     = '#0F172A'   # Title bg          (Slate 900)
    C_H2     = '#1E293B'   # Subtitle bg        (Slate 800)
    C_HDR    = '#334155'   # Col header bg      (Slate 700)
    C_H1_FG  = '#FFFFFF'   # Title text
    C_H2_FG  = '#F1F5F9'   # Subtitle text      (Slate 100)
    C_DATE   = '#FFFFFF'   # Date numbers       (white — highest contrast)
    C_WDAY   = '#F1F5F9'   # Weekday abbrevs    (Slate 100)
    C_HDR_FG = '#FFFFFF'   # Col hdr text
    # Sunday column header — Slate Crimson
    C_SH_BG  = '#991B1B'   # Red 800
    C_SH_FG  = '#FFFFFF'   # White
    # DATA ROWS
    C_ODD    = '#FFFFFF'   # White
    C_EVEN   = '#F8FAFC'   # Slate 50 — clean silver-slate wash
    # FIXED LEFT COLUMNS
    C_LC_BG  = '#F8FAFC'   # Slate 50
    C_LC_SEP = '#CBD5E1'   # Slate 200 — left-col border
    # TEXT
    C_T_DARK = '#0F172A'   # Slate 900
    C_T_MED  = '#334155'   # Slate 700
    C_T_MUTE = '#64748B'   # Slate 500
    # BORDERS
    C_B_OUT  = '#0F172A'   # Outer boundary
    C_B_IN   = '#CBD5E1'   # Inner border (Slate 200)
    C_B_LC   = '#CBD5E1'   # Left-col separator
    #
    # STATUS COLORS — Tailwind 100/800 pairs
    #
    S_R_BG  = '#F1F5F9'; S_R_FG  = '#475569'  # Rest (Slate 100 / Slate 600)
    S_CR_BG = '#D1FAE5'; S_CR_FG = '#065F46'  # CR (Emerald 100 / Emerald 800)
    S_PN_BG = '#F5F3FF'; S_PN_FG = '#6D28D9'  # P/N (Violet 50 / Violet 700)
    S_CL_BG = '#DBEAFE'; S_CL_FG = '#1E40AF'  # CL (Blue 100 / Blue 800)
    S_LA_BG = '#FCE7F3'; S_LA_FG = '#9D174D'  # LAP (Pink 100 / Pink 800)
    S_SK_BG = '#FEE2E2'; S_SK_FG = '#991B1B'  # Sick (Red 100 / Red 800)
    S_SC_BG = '#EDE9FE'; S_SC_FG = '#5B21B6'  # SCL (Purple 100 / Purple 800)
    S_PH_BG = '#FEF9C3'; S_PH_FG = '#713F12'  # PH (Yellow 100 / Yellow 900)
    S_CU_BG = '#E0F2FE'; S_CU_FG = '#0369A1'  # Custom (Sky 100 / Sky 800)
    S_SN_BG = '#FEE2E2'; S_SN_FG = '#991B1B'  # Sunday data cell
    S_HL_BG = '#FEF9C3'; S_HL_FG = '#713F12'  # Holiday data cell
    # Section banner
    C_SEC_BG = '#1E293B'; C_SEC_FG = '#F1F5F9'

    # ── Format factory ────────────────────────────────────────────────────────
    def F(props):
        base = {'font_name': FONT, 'valign': 'vcenter'}
        base.update(props)
        return workbook.add_format(base)

    # ── Static formats ────────────────────────────────────────────────────────

    f_title = F({
        'font_size': 13, 'bold': True, 'align': 'center',
        'bg_color': C_H1, 'font_color': C_H1_FG,
        'top': 5, 'top_color': C_H1,
        'left': 5, 'left_color': C_H1,
        'right': 5, 'right_color': C_H1,
    })
    f_subtitle = F({
        'font_size': 9, 'align': 'center',
        'bg_color': C_H2, 'font_color': C_H2_FG,
        'left': 5, 'left_color': C_H1,
        'right': 5, 'right_color': C_H1,
        'bottom': 2, 'bottom_color': '#475569',
    })
    f_col_hdr = F({
        'font_size': 8.5, 'bold': True, 'align': 'center',
        'bg_color': C_HDR, 'font_color': C_HDR_FG,
        'border': 1, 'border_color': '#1E293B',
        'text_wrap': True,
    })
    # Date row — big, white, bold
    f_date = F({
        'font_size': 10, 'bold': True, 'align': 'center',
        'bg_color': C_HDR, 'font_color': C_DATE,
        'border': 1, 'border_color': '#1E293B',
    })
    # Weekday row — comfortable size, light gray text (clearly visible)
    f_wday = F({
        'font_size': 8, 'bold': False, 'align': 'center',
        'bg_color': C_HDR, 'font_color': C_WDAY,
        'border': 1, 'border_color': '#1E293B',
    })
    # Sunday column header
    f_sun_date = F({
        'font_size': 10, 'bold': True, 'align': 'center',
        'bg_color': C_SH_BG, 'font_color': C_SH_FG,
        'border': 1, 'border_color': '#7F1D1D',
    })
    f_sun_wday = F({
        'font_size': 8, 'bold': True, 'align': 'center',
        'bg_color': C_SH_BG, 'font_color': C_SH_FG,
        'border': 1, 'border_color': '#7F1D1D',
    })

    # Fixed left-column formats (same every row)
    f_sl = F({
        'font_size': 8.5, 'bold': True, 'align': 'center',
        'bg_color': C_LC_BG, 'font_color': C_T_MUTE,
        'border': 1, 'border_color': C_B_LC,
        'left': 3, 'left_color': C_B_OUT,
    })
    f_name = F({
        'font_size': 9, 'bold': True, 'align': 'left',
        'bg_color': C_LC_BG, 'font_color': C_T_DARK,
        'border': 1, 'border_color': C_B_LC,
        'text_wrap': True,
    })
    f_desig = F({
        'font_size': 8.5, 'italic': True, 'align': 'center',
        'bg_color': C_LC_BG, 'font_color': C_T_MED,
        'border': 1, 'border_color': C_B_LC,
        'right': 2, 'right_color': C_HDR,
    })

    # Per-row factories
    def f_empty(bg):
        return F({'font_size': 8.5, 'align': 'center',
                  'bg_color': bg, 'font_color': C_T_MUTE,
                  'border': 1, 'border_color': C_B_IN})
    # Present cell — clean, minimal (dark blue for high contrast)
    def f_p(bg):
        return F({'font_size': 8.5, 'bold': True, 'align': 'center',
                  'bg_color': bg, 'font_color': '#1E40AF',
                  'border': 1, 'border_color': C_B_IN})
    def f_remarks(bg):
        return F({'font_size': 8.5, 'align': 'left',
                  'bg_color': bg, 'font_color': C_T_MUTE,
                  'border': 1, 'border_color': C_B_IN,
                  'right': 3, 'right_color': C_B_OUT,
                  'text_wrap': True})

    # Status cell factory
    def sf(bg, fg, bold=True, italic=False):
        return F({'font_size': 8.5, 'bold': bold, 'italic': italic,
                  'align': 'center', 'bg_color': bg, 'font_color': fg,
                  'border': 1, 'border_color': C_B_IN,
                  'text_wrap': True})

    # Pre-built status formats
    f_r    = sf(S_R_BG,  S_R_FG,  bold=False, italic=False)  # Rest — plain
    f_cr   = sf(S_CR_BG, S_CR_FG, bold=True,  italic=False)  # CR — bold
    f_pn   = sf(S_PN_BG, S_PN_FG, bold=True,  italic=False)  # P/N — bold
    f_cl   = sf(S_CL_BG, S_CL_FG, bold=True,  italic=False)  # CL — bold
    f_lap  = sf(S_LA_BG, S_LA_FG, bold=True,  italic=True)   # LAP — bold+italic
    f_sick = sf(S_SK_BG, S_SK_FG, bold=True,  italic=True)   # Sick — bold+italic
    f_scl  = sf(S_SC_BG, S_SC_FG, bold=True,  italic=False)  # SCL — bold
    f_ph   = sf(S_PH_BG, S_PH_FG, bold=True,  italic=False)  # PH — bold
    f_cu   = sf(S_CU_BG, S_CU_FG, bold=True,  italic=True)   # Custom — bold+italic
    f_sun  = sf(S_SN_BG, S_SN_FG, bold=True,  italic=False)  # Sunday cell
    f_hol  = sf(S_HL_BG, S_HL_FG, bold=True,  italic=False)  # Holiday cell

    # CR rich-text sub-formats
    f_cr_main = F({'font_size': 8.5, 'bold': True,  'font_color': S_CR_FG, 'bg_color': S_CR_BG})
    f_cr_sub  = F({'font_size': 6.5, 'bold': False, 'font_color': '#3B82F6', 'bg_color': S_CR_BG})

    # Section banner (joint ALL view)
    f_sec = F({'font_size': 9, 'bold': True, 'align': 'left',
               'bg_color': C_SEC_BG, 'font_color': C_SEC_FG,
               'border': 1, 'border_color': C_B_OUT})

    # Footer
    f_lbl   = F({'font_size': 8, 'italic': True, 'align': 'center', 'font_color': C_T_MUTE})
    f_sig_l = F({'font_size': 9, 'bold': True, 'align': 'left',
                 'valign': 'top', 'text_wrap': True, 'font_color': C_T_DARK})
    f_sig_r = F({'font_size': 9, 'bold': True, 'align': 'right',
                 'valign': 'top', 'text_wrap': True, 'font_color': C_T_DARK})

    # ── Header block (dynamic range based on RC) ───────────────────────────────
    sheet.set_row(0, 26)
    sheet.set_row(1, 20)
    sheet.set_row(2, 20)
    sheet.set_row(3, 16)

    sheet.merge_range(0, 0, 0, LC, "METRO RAILWAY, KOLKATA", f_title)
    sheet.merge_range(
        1, 0, 1, LC,
        f"Attendance Sheet  \u00b7  Signal Department  ({req.section_name})"
        f"     Period:  {req.period_start}  \u2013  {req.period_end}",
        f_subtitle
    )

    # ── Column header rows ────────────────────────────────────────────────────
    sheet.merge_range(2, 0, 3, 0, "SL",                      f_col_hdr)
    sheet.merge_range(2, 1, 3, 1, "Name of Staff\nP.F. No.", f_col_hdr)
    sheet.merge_range(2, 2, 3, 2, "Designation",             f_col_hdr)

    for d_idx, day_obj in enumerate(days_in_grid):
        cc = 3 + d_idx
        if day_obj.weekday == "Sun":
            sheet.write(2, cc, day_obj.day,     f_sun_date)
            sheet.write(3, cc, day_obj.weekday, f_sun_wday)
        else:
            sheet.write(2, cc, day_obj.day,     f_date)
            sheet.write(3, cc, day_obj.weekday, f_wday)

    sheet.merge_range(2, RC, 3, RC, "Remarks", f_col_hdr)

    # ── Data rows ─────────────────────────────────────────────────────────────
    curr_row = 4
    is_joint = req.section_code == 'ALL'
    from collections import defaultdict
    grouped = defaultdict(list)
    for emp in req.rows:
        grouped[emp.section_code or 'KKVS'].append(emp)

    emp_no = 0
    for sec in sorted(grouped.keys()):
        emps = grouped[sec]

        if is_joint:
            label = f"  \u25b8  {sec.upper()}   \u2014   {get_section_name_from_db(sec)}"
            sheet.merge_range(curr_row, 0, curr_row, LC, label, f_sec)
            sheet.set_row(curr_row, 20)
            curr_row += 1

        for emp in emps:
            emp_no += 1
            bg = C_ODD if emp_no % 2 == 1 else C_EVEN
            sheet.set_row(curr_row, 36) # dynamic and tall row height

            sheet.write(curr_row, 0, emp.sl,           f_sl)
            sheet.write(curr_row, 1, f"{emp.name}\n(PF: {emp.pf_number})", f_name)
            sheet.write(curr_row, 2, emp.designation,  f_desig)

            # Consecutive-status spans
            spans, n, i = [], len(emp.days), 0
            while i < n:
                st = emp.days[i].status
                s = i
                while i < n - 1 and emp.days[i + 1].status == st:
                    i += 1
                spans.append((s, i, st))
                i += 1

            STANDARD_NON_MERGED = {'P', 'P/N', 'R', 'CR', 'CL', 'LAP', 'SCL', 'PH', '', None}

            for si, ei, val in spans:
                sc = 3 + si
                ec = 3 + ei

                if val in STANDARD_NON_MERGED:
                    for ci in range(sc, ec + 1):
                        dobj = emp.days[ci - 3]
                        if val == 'CR':
                            done = False
                            if dobj.remarks and "CR_EARNED_DATE:" in dobj.remarks:
                                edate = dobj.remarks.split("CR_EARNED_DATE:")[1].strip()[:10]
                                try:
                                    dt = datetime.strptime(edate, "%Y-%m-%d")
                                    sheet.write_rich_string(
                                        curr_row, ci,
                                        f_cr_main, 'CR\n',
                                        f_cr_sub,  dt.strftime('%d.%m'),
                                        f_cr)
                                    done = True
                                except Exception:
                                    pass
                            if not done:
                                sheet.write(curr_row, ci, 'CR', f_cr)
                        elif val == 'P':
                            if dobj.weekday == 'Sun':
                                sheet.write(curr_row, ci, 'P', f_sun)
                            elif dobj.is_holiday:
                                sheet.write(curr_row, ci, 'P', f_hol)
                            else:
                                sheet.write(curr_row, ci, 'P', f_p(bg))
                        elif val == 'P/N':
                            sheet.write(curr_row, ci, 'P/N', f_pn)
                        elif val == 'R':
                            sheet.write(curr_row, ci, 'R', f_r)
                        elif val == 'CL':
                            sheet.write(curr_row, ci, 'CL', f_cl)
                        elif val == 'LAP':
                            sheet.write(curr_row, ci, 'LAP', f_lap)
                        elif val == 'SCL':
                            sheet.write(curr_row, ci, 'SCL', f_scl)
                        elif val == 'PH':
                            sheet.write(curr_row, ci, 'PH', f_ph)
                        else:
                            if dobj.weekday == 'Sun':
                                sheet.write(curr_row, ci, '', f_sun)
                            elif dobj.is_holiday:
                                sheet.write(curr_row, ci, '', f_hol)
                            else:
                                sheet.write(curr_row, ci, '', f_empty(bg))
                else:
                    order = emp.days[si].remarks
                    if val == 'Sick':
                        if sc == ec:
                            sheet.write(curr_row, sc, 'Sick', f_sick)
                        else:
                            sheet.merge_range(curr_row, sc, curr_row, ec, 'Sick', f_sick)
                    else:
                        if sc == ec:
                            disp = val
                            if order:
                                disp = f"{val}\n({order})"
                            sheet.write(curr_row, sc, disp, f_cu)
                        else:
                            disp = f"◀  {val}  ▶"
                            if order:
                                disp = f"◀  {val} ({order})  ▶"
                            sheet.merge_range(curr_row, sc, curr_row, ec, disp, f_cu)

            sheet.write(curr_row, RC, emp.remarks or '', f_remarks(bg))
            curr_row += 1

    # ── Footer ────────────────────────────────────────────────────────────────
    sheet.set_row(curr_row, 10)      # spacer
    curr_row += 1

    rule_f = F({'bg_color': C_HDR, 'border': 0})
    sheet.merge_range(curr_row, 0, curr_row, LC, '', rule_f)
    sheet.set_row(curr_row, 2)
    curr_row += 1

    sheet.merge_range(
        curr_row, 0, curr_row, LC,
        f"Prepared on: {req.submission_date}   \u00b7   "
        f"Signal Department, Metro Railway Kolkata   \u00b7   {req.section_name}",
        f_lbl)
    sheet.set_row(curr_row, 15)
    curr_row += 2

    # Signature row with dynamic safe height (70pt)
    sig_row = curr_row
    sheet.set_row(sig_row, 70)

    left_end  = max(13, RC // 2 - 2)
    right_start = left_end + 3
    sheet.merge_range(
        sig_row, 0, sig_row, left_end,
        f"{req.signatory_left}\n(Prepared & Submitted by)", f_sig_l)
    sheet.merge_range(
        sig_row, right_start, sig_row, LC,
        f"For {req.signatory_right}\n(Verified / Countersigned)", f_sig_r)

    workbook.close()
    output.seek(0)

    filename = f"Attendance_Sheet_{req.section_code}_{req.period_start}_{req.period_end}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )

# --- Endpoint 4: Export Attendance PDF ---
@app.post("/api/export/attendance/pdf")
async def export_attendance_pdf(req: AttendanceExportRequest):
    pdf_buffer = io.BytesIO()
    
    # Needs landscape and centered margins to fit 31 columns of attendance
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        rightMargin=56,
        leftMargin=56,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'AttTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        alignment=1,
        textColor=colors.HexColor("#1B365D")
    )
    
    meta_style = ParagraphStyle(
        'AttMeta',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        alignment=1
    )

    cell_style = ParagraphStyle(
        'AttCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=5.5,
        leading=6.5,
        alignment=1,
        textColor=colors.HexColor('#0F172A')
    )
    
    cell_bold_style = ParagraphStyle(
        'AttCellBold',
        parent=cell_style,
        fontName='Helvetica-Bold',
        fontSize=5.5,
        leading=6.5,
        textColor=colors.HexColor('#0F172A')
    )

    day_header_style = ParagraphStyle(
        'AttDayHeader',
        parent=cell_bold_style,
        fontSize=5.5,
        leading=6
    )

    att_header_text_style = ParagraphStyle(
        'AttTableHeaderText',
        parent=cell_bold_style,
        textColor=colors.white
    )

    att_day_header_text_style = ParagraphStyle(
        'AttTableDayHeaderText',
        parent=day_header_style,
        textColor=colors.white
    )

    name_style = ParagraphStyle(
        'AttName',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=8
    )

    name_pf_style = ParagraphStyle(
        'AttNamePF',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=8.5,
        alignment=1 # Center
    )

    story = []

    story.append(Paragraph("METRO RAILWAY, KOLKATA", title_style))
    story.append(Paragraph(f"ATTENDANCE SHEET OF SIGNAL DEPARTMENT ({req.section_name.upper()}) — For the Period from {req.period_start} to {req.period_end}", ParagraphStyle('AttSub', parent=title_style, fontSize=10)))
    story.append(Spacer(1, 10))

    # Table headers (Name and PF combined in column 2)
    header_row_1 = [
        Paragraph("<b>SL</b>", att_header_text_style),
        Paragraph("<b>Name of Staff / P.F. No.</b>", att_header_text_style),
        Paragraph("<b>Designation</b>", att_header_text_style)
    ]
    
    days_in_grid = req.rows[0].days if req.rows else []
    for day in days_in_grid:
        header_row_1.append(Paragraph(f"<b>{day.day}</b><br/>{day.weekday[0]}", att_day_header_text_style))
        
    header_row_1.append(Paragraph("<b>Remarks</b>", att_header_text_style))

    table_data = [header_row_1]

    # Construct grid styles, highlight sundays in table style backgrounds
    t_style = [
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1B365D")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (1,1), (1,-1), 'CENTER'), # Center Name/PF
        ('BOTTOMPADDING', (0,0), (-1,-1), 1.5),
        ('TOPPADDING', (0,0), (-1,-1), 1.5),
        ('LEFTPADDING', (0,0), (-1,-1), 1),
        ('RIGHTPADDING', (0,0), (-1,-1), 1),
    ]

    is_joint_view = req.section_code == 'ALL'
    from collections import defaultdict
    grouped_rows = defaultdict(list)
    for row in req.rows:
        grouped_rows[row.section_code or 'KKVS'].append(row)
        
    section_title_style = ParagraphStyle(
        'PDFAttSectionTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#1E293B")
    )
    
    row_num = 1 # Index 0 is header_row_1
    for sec_code in sorted(grouped_rows.keys()):
        rows_in_sec = grouped_rows[sec_code]
        
        if is_joint_view:
            banner_text = f"<b>SECTION: {sec_code.upper()}</b> ({get_section_name_from_db(sec_code)})"
            banner_row = [Paragraph(banner_text, section_title_style)] + [""] * (len(days_in_grid) + 2)
            table_data.append(banner_row)
            
            t_style.append(('SPAN', (0, row_num), (-1, row_num)))
            t_style.append(('BACKGROUND', (0, row_num), (-1, row_num), colors.HexColor("#E2E8F0")))
            t_style.append(('ALIGN', (0, row_num), (-1, row_num), 'LEFT'))
            t_style.append(('BOTTOMPADDING', (0, row_num), (-1, row_num), 3))
            t_style.append(('TOPPADDING', (0, row_num), (-1, row_num), 3))
            row_num += 1
            
        for idx, row in enumerate(rows_in_sec):
            row_bg = colors.HexColor("#F8FAFC") if idx % 2 == 1 else colors.white
            t_style.append(('BACKGROUND', (0, row_num), (2, row_num), row_bg))
            t_style.append(('BACKGROUND', (-1, row_num), (-1, row_num), row_bg))
            
            name_pf_html = f"<b>{row.name}</b><br/><font size=5.5 color='#4A5568'>PF: {row.pf_number}</font>"
            
            data_line = [
                Paragraph(str(row.sl), cell_style),
                Paragraph(name_pf_html, name_pf_style),
                Paragraph(row.designation, cell_style)
            ]
            
            # Spans logic
            spans = []
            n = len(row.days)
            i = 0
            while i < n:
                status = row.days[i].status
                start_idx = i
                while i < n - 1 and row.days[i+1].status == status:
                    i += 1
                end_idx = i
                spans.append((start_idx, end_idx, status))
                i += 1
                
            day_cells = [None] * len(days_in_grid)
            for start_idx, end_idx, status in spans:
                should_merge = status not in ('P', 'CR', '', None) and start_idx < end_idx
                
                if should_merge:
                    order = row.days[start_idx].remarks
                    if status not in ('CL', 'LAP', 'Sick', 'SCL', 'PH', 'R', 'P/N') and order:
                        status_html = f"<b>{status}</b><br/><font size=4>{order}</font>"
                    else:
                        status_html = f"<b>{status}</b>"
                    day_cells[start_idx] = Paragraph(status_html, cell_bold_style)
                    for s_idx in range(start_idx + 1, end_idx + 1):
                        day_cells[s_idx] = ""
                    t_style.append(('SPAN', (3 + start_idx, row_num), (3 + end_idx, row_num)))
                else:
                    for s_idx in range(start_idx, end_idx + 1):
                        val = row.days[s_idx].status
                        day_obj = row.days[s_idx]
                        
                        display_html = val
                        if val == 'CR' and day_obj.remarks:
                            if "CR_EARNED_DATE:" in day_obj.remarks:
                                edate = day_obj.remarks.split("CR_EARNED_DATE:")[1].strip()[:10]
                                try:
                                    dt = datetime.strptime(edate, "%Y-%m-%d")
                                    display_html = f"<b>CR</b><br/><font size=4 color='#1D4ED8'>{dt.strftime('%d.%m')}</font>"
                                except Exception:
                                    display_html = f"<b>{val}</b>"
                            else:
                                display_html = f"<b>{val}</b>"
                        elif val in ('R', 'Sick', 'CL', 'LAP', 'SCL', 'PH', 'P/N'):
                            display_html = f"<b>{val}</b>"
                        elif val not in ('P', '', None):
                            if day_obj.remarks:
                                display_html = f"<b>{val}</b><br/><font size=4>{day_obj.remarks}</font>"
                            else:
                                display_html = f"<b>{val}</b>"
                            
                        day_cells[s_idx] = Paragraph(display_html, cell_bold_style if val not in ('P', '', None) else cell_style)
                        
            for val in day_cells:
                data_line.append(val)
                
            data_line.append(Paragraph(row.remarks or "", cell_style))
            table_data.append(data_line)

            # Apply coloring to individual/span cells
            for d_idx, day in enumerate(row.days):
                col_c = 3 + d_idx
                val = day.status
                color_map = {
                    'P/N': '#F3E8FF',
                    'R': '#F1F5F9',
                    'CR': '#EFF6FF',
                    'CL': '#FFFBEB',
                    'LAP': '#FFF7ED',
                    'Sick': '#FEF2F2',
                    'SCL': '#FFF1F2',
                    'PH': '#FEF9C3',
                }
                if val in color_map:
                    t_style.append(('BACKGROUND', (col_c, row_num), (col_c, row_num), colors.HexColor(color_map[val])))
                elif val not in ('P', '', None):
                    t_style.append(('BACKGROUND', (col_c, row_num), (col_c, row_num), colors.HexColor('#E6FFFA')))
                else:
                    if day.weekday == "Sun":
                        t_style.append(('BACKGROUND', (col_c, row_num), (col_c, row_num), colors.HexColor("#FEE2E2")))
                    elif day.is_holiday:
                        t_style.append(('BACKGROUND', (col_c, row_num), (col_c, row_num), colors.HexColor("#FEF3C7")))
                        
            row_num += 1

    # Printable area centered with margins
    # Total width for landscape A4 = 841.89 - 56*2 = 729.89pt
    # col widths safely within boundaries to avoid page overflow:
    col_widths = [14, 90, 60] + [15] * len(days_in_grid) + [45]
    total_table_width = 14 + 90 + 60 + (15 * len(days_in_grid)) + 45
    
    grid_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    grid_table.setStyle(TableStyle(t_style))
    story.append(grid_table)
    story.append(Spacer(1, 15))
 
    # Signatures aligned dynamically to match total grid table width
    sig_left_text = req.signatory_left.replace('\n', '<br/>')
    sig_right_text = f"For {req.signatory_right}".replace('\n', '<br/>')
    
    sig_data = [
        [Paragraph(f"<b>{sig_left_text}</b>", ParagraphStyle('LeftSigP', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, alignment=0)),
         Paragraph(f"<b>{sig_right_text}</b>", ParagraphStyle('RightSigP', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, alignment=2))],
        [Paragraph("S&T Dept., Metro Railway, Kolkata", ParagraphStyle('LeftSigSub', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, leading=9, alignment=0, textColor=colors.HexColor("#4A5568"))),
         Paragraph("Metro Railway, Kolkata", ParagraphStyle('RightSigSub', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, leading=9, alignment=2, textColor=colors.HexColor("#4A5568")))]
    ]
    sig_col_width = total_table_width / 2.0
    sig_table = Table(sig_data, colWidths=[sig_col_width, sig_col_width])
    sig_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ('TOPPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(sig_table)

    doc.build(story, canvasmaker=NumberedCanvas)
    pdf_buffer.seek(0)
    
    filename = f"Attendance_Sheet_{req.section_code}_{req.period_start}_{req.period_end}.pdf"
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )

# --- Traveling Allowance Pydantic Schemas & Endpoints ---
class TAEntrySchema(BaseModel):
    id: Optional[int] = None
    entry_date: str
    train_no: Optional[str] = ""
    time_left: Optional[str] = ""
    time_arrived: Optional[str] = ""
    station_from: Optional[str] = ""
    station_to: Optional[str] = ""
    is_stay: Optional[int] = 0
    stay_details: Optional[str] = ""
    days_nights: Optional[str] = ""
    object_journey: Optional[str] = ""
    rate: int
    amount: int

class TABillSchema(BaseModel):
    id: Optional[int] = None
    emp_id: int
    month_year: str
    journey_type: str
    book_no: Optional[str] = ""
    page_no: Optional[str] = ""
    serial_no_from: Optional[str] = ""
    serial_no_to: Optional[str] = ""
    bill_unit: Optional[str] = ""
    basic_pay: Optional[int] = 0
    entries: List[TAEntrySchema]

@app.get("/api/ta-bills")
def get_ta_bills(section_code: Optional[str] = None):
    conn = get_db()
    query = """
        SELECT b.*, e.name as emp_name, e.pf_number, e.designation, e.level
        FROM ta_bills b
        JOIN employees e ON b.emp_id = e.emp_id
        LEFT JOIN sections s ON e.primary_section_id = s.id
    """
    params = []
    if section_code and section_code != 'ALL':
        query += " WHERE s.section_code = ?"
        params.append(section_code)
    
    query += " ORDER BY b.month_year DESC, b.id DESC"
    rows = conn.execute(query, params).fetchall()
    
    bills = []
    for row in rows:
        b = dict(row)
        total_amount = conn.execute("SELECT SUM(amount) FROM ta_entries WHERE bill_id = ?", (b['id'],)).fetchone()[0] or 0
        b['total_amount'] = total_amount
        bills.append(b)
        
    conn.close()
    return bills

@app.get("/api/ta-bills/{id}")
def get_ta_bill(id: int):
    conn = get_db()
    row = conn.execute("""
        SELECT b.*, e.name as emp_name, e.pf_number, e.designation, e.level
        FROM ta_bills b
        JOIN employees e ON b.emp_id = e.emp_id
        WHERE b.id = ?
    """, (id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="TA Bill not found")
    
    bill = dict(row)
    entries = [dict(r) for r in conn.execute("SELECT * FROM ta_entries WHERE bill_id = ? ORDER BY id ASC", (id,)).fetchall()]
    bill['entries'] = entries
    conn.close()
    return bill

@app.post("/api/ta-bills")
def create_ta_bill(payload: TABillSchema):
    conn = get_db()
    cursor = conn.cursor()
    try:
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("""
            INSERT INTO ta_bills (emp_id, month_year, journey_type, book_no, page_no, serial_no_from, serial_no_to, bill_unit, basic_pay, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (payload.emp_id, payload.month_year, payload.journey_type, payload.book_no, payload.page_no, payload.serial_no_from, payload.serial_no_to, payload.bill_unit or "", payload.basic_pay or 0, now_str))
        bill_id = cursor.lastrowid
        
        for entry in payload.entries:
            cursor.execute("""
                INSERT INTO ta_entries (bill_id, entry_date, train_no, time_left, time_arrived, station_from, station_to, is_stay, stay_details, days_nights, object_journey, rate, amount)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (bill_id, entry.entry_date, entry.train_no, entry.time_left, entry.time_arrived, entry.station_from, entry.station_to, entry.is_stay, entry.stay_details, entry.days_nights, entry.object_journey, entry.rate, entry.amount))
            
        conn.commit()
        log_audit("Insert", "TA Bills", f"Created TA Bill ID {bill_id} for Employee ID {payload.emp_id}")
        return {"id": bill_id, "status": "success"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
 
@app.put("/api/ta-bills/{id}")
def update_ta_bill(id: int, payload: TABillSchema):
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            UPDATE ta_bills 
            SET emp_id = ?, month_year = ?, journey_type = ?, book_no = ?, page_no = ?, serial_no_from = ?, serial_no_to = ?, bill_unit = ?, basic_pay = ?
            WHERE id = ?
        """, (payload.emp_id, payload.month_year, payload.journey_type, payload.book_no, payload.page_no, payload.serial_no_from, payload.serial_no_to, payload.bill_unit or "", payload.basic_pay or 0, id))
        
        cursor.execute("DELETE FROM ta_entries WHERE bill_id = ?", (id,))
        for entry in payload.entries:
            cursor.execute("""
                INSERT INTO ta_entries (bill_id, entry_date, train_no, time_left, time_arrived, station_from, station_to, is_stay, stay_details, days_nights, object_journey, rate, amount)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (id, entry.entry_date, entry.train_no, entry.time_left, entry.time_arrived, entry.station_from, entry.station_to, entry.is_stay, entry.stay_details, entry.days_nights, entry.object_journey, entry.rate, entry.amount))
            
        conn.commit()
        log_audit("Update", "TA Bills", f"Updated TA Bill ID {id} for Employee ID {payload.emp_id}")
        return {"id": id, "status": "success"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.delete("/api/ta-bills/{id}")
def delete_ta_bill(id: int):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM ta_bills WHERE id = ?", (id,))
        conn.commit()
        log_audit("Delete", "TA Bills", f"Deleted TA Bill ID {id}")
        return {"status": "success"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

def num_to_words(n):
    n = int(round(n))
    if n == 0:
        return "Zero"
    
    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten", 
             "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    
    def helper(num):
        if num < 20:
            return units[num]
        elif num < 100:
            return tens[num // 10] + (" " + units[num % 10] if num % 10 != 0 else "")
        elif num < 1000:
            return units[num // 100] + " Hundred" + (" and " + helper(num % 100) if num % 100 != 0 else "")
        elif num < 100000:
            return helper(num // 1000) + " Thousand" + (" " + helper(num % 1000) if num % 1000 != 0 else "")
        else:
            return helper(num // 100000) + " Lakh" + (" " + helper(num % 100000) if num % 100000 != 0 else "")
            
    res = helper(n)
    return res

def delete_rows_clean(ws, start_row, amount):
    import openpyxl
    new_ranges = []
    for r in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = r.bounds
        if min_row >= start_row and max_row < start_row + amount:
            continue
        if min_row >= start_row + amount:
            r.shift(row_shift=-amount)
            new_ranges.append(r)
        elif max_row < start_row:
            new_ranges.append(r)
        else:
            new_max_row = max_row
            if max_row >= start_row:
                if max_row < start_row + amount:
                    new_max_row = start_row - 1
                else:
                    new_max_row = max_row - amount
            if new_max_row >= min_row:
                r.max_row = new_max_row
                new_ranges.append(r)
    ws.merged_cells.ranges.clear()
    for r in new_ranges:
        ws.merged_cells.ranges.add(r)
    ws.delete_rows(start_row, amount)

def insert_rows_clean(ws, start_row, amount):
    import openpyxl
    new_ranges = []
    for r in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = r.bounds
        if min_row >= start_row:
            r.shift(row_shift=amount)
            new_ranges.append(r)
        elif max_row < start_row:
            new_ranges.append(r)
        else:
            r.max_row = max_row + amount
            new_ranges.append(r)
    ws.merged_cells.ranges.clear()
    for r in new_ranges:
        ws.merged_cells.ranges.add(r)
    ws.insert_rows(start_row, amount)

def copy_row_style(ws, src_row, dst_row):
    import openpyxl
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell.font = openpyxl.styles.Font(
                name=src_cell.font.name,
                size=src_cell.font.size,
                bold=src_cell.font.bold,
                italic=src_cell.font.italic,
                charset=src_cell.font.charset,
                color=src_cell.font.color,
                underline=src_cell.font.underline,
                strike=src_cell.font.strike,
                vertAlign=src_cell.font.vertAlign,
                scheme=src_cell.font.scheme
            )
            dst_cell.border = openpyxl.styles.Border(
                left=src_cell.border.left,
                right=src_cell.border.right,
                top=src_cell.border.top,
                bottom=src_cell.border.bottom
            )
            dst_cell.fill = openpyxl.styles.PatternFill(
                fill_type=src_cell.fill.fill_type,
                start_color=src_cell.fill.start_color,
                end_color=src_cell.fill.end_color
            )
            dst_cell.alignment = openpyxl.styles.Alignment(
                horizontal=src_cell.alignment.horizontal,
                vertical=src_cell.alignment.vertical,
                text_rotation=src_cell.alignment.text_rotation,
                wrap_text=src_cell.alignment.wrap_text,
                shrink_to_fit=src_cell.alignment.shrink_to_fit,
                indent=src_cell.alignment.indent
            )
            dst_cell.number_format = src_cell.number_format

@app.get("/api/ta-bills/{id}/export-excel")
def export_ta_bill_excel(id: int):
    import openpyxl
    import re
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    def apply_premium_styling(ws, journey_type, required_rows):
        # Fonts
        title_font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
        meta_font = Font(name="Segoe UI", size=9.5, bold=False, color="333333")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=9.5, color="222222")
        data_bold_font = Font(name="Segoe UI", size=9.5, bold=True, color="222222")
        total_font = Font(name="Segoe UI", size=10.5, bold=True, color="1B365D")
        
        # Fills
        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        
        # Borders
        thin_border_side = Side(border_style="thin", color="CBD5E1")
        thick_border_side = Side(border_style="medium", color="1B365D")
        double_border_side = Side(border_style="double", color="1B365D")
        
        data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        header_border = Border(left=thin_border_side, right=thin_border_side, top=thick_border_side, bottom=thick_border_side)
        total_border = Border(top=thin_border_side, bottom=double_border_side, left=thin_border_side, right=thin_border_side)

        # Column count
        max_col = 10 if journey_type == "NORMAL" else 11
        
        # Row Heights
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 24
        ws.row_dimensions[4].height = 24
        ws.row_dimensions[5].height = 24
        ws.row_dimensions[7].height = 28
        ws.row_dimensions[8].height = 20

        # Helper to check if a cell is merged across columns
        def is_merged_across_columns(ws, row, col):
            for r_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = r_range.bounds
                if min_row <= row <= max_row and min_col <= col <= max_col:
                    if min_col != max_col:
                        return True
            return False

        # Helper to estimate needed height for wrapped text
        def calculate_needed_row_height(text, col_width, font_size=9.5, bold=False):
            if not text:
                return 20
            char_width = 0.85 if not bold else 0.95
            words = str(text).split(" ")
            lines = []
            current_line = ""
            for word in words:
                if "\n" in word:
                    parts = word.split("\n")
                    for idx, part in enumerate(parts):
                        if idx == 0:
                            if len(current_line + " " + part) * char_width <= col_width:
                                current_line = (current_line + " " + part).strip()
                            else:
                                lines.append(current_line)
                                current_line = part.strip()
                        else:
                            lines.append(current_line)
                            current_line = part.strip()
                else:
                    if len(current_line + " " + word) * char_width <= col_width:
                        current_line = (current_line + " " + word).strip()
                    else:
                        lines.append(current_line)
                        current_line = word.strip()
            if current_line:
                lines.append(current_line)
            num_lines = max(1, len(lines))
            line_height = font_size * 1.35
            return num_lines * line_height + 10  # height in points with padding

        # Base/minimum widths for all columns to keep it balanced and prevent header clipping
        if journey_type == "NORMAL":
            base_widths = {
                1: 13,  # Month & Date
                2: 18,  # No. of Train / Steamer / Plain
                3: 13,  # Time left (Hrs.)
                4: 13,  # Time arrived (Hrs.)
                5: 15,  # Station From
                6: 15,  # Station To
                7: 13,  # Days / Nights
                8: 45,  # Object of journey
                9: 14,  # Rate in Rs.
                10: 16  # Amount in Rs.
            }
        else:
            base_widths = {
                1: 13,  # Month & Date
                2: 18,  # No. of Train / Steamer / Plain
                3: 13,  # Time left (Hrs.)
                4: 13,  # Time arrived (Hrs.)
                5: 15,  # Station From
                6: 15,  # Station To
                7: 14,  # More than 8 KMs
                8: 13,  # Days / Nights
                9: 45,  # Object of journey
                10: 14, # Rate in Rs.
                11: 16  # Amount in Rs.
            }
            
        data_start = 9
        data_end = 8 + required_rows
        total_row = 9 + required_rows

        # Calculate optimal width for columns 1 to max_col based on rows 7 to total_row (table area)
        for c in range(1, max_col + 1):
            max_len = base_widths.get(c, 12)
            for r in range(7, total_row + 1):
                if is_merged_across_columns(ws, r, c):
                    continue
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    val_str = str(val)
                    lines = val_str.split("\n")
                    for line in lines:
                        max_len = max(max_len, len(line) + 3)
            
            limit = 55 if c == (8 if journey_type == "NORMAL" else 9) else 35
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = min(max_len, limit)

        # 2. Format Headers (Rows 7 and 8)
        for r in [7, 8]:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = header_border

        # 3. Format Data Rows (Row 9 to 8+required_rows)
        for r in range(data_start, data_end + 1):
            if journey_type == "NORMAL":
                pair_idx = (r - data_start) // 2
                fill_to_use = alt_fill if pair_idx % 2 == 1 else white_fill
            else:
                fill_to_use = alt_fill if (r % 2 == 1) else white_fill
                
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = data_font
                cell.fill = fill_to_use
                cell.border = data_border
                
                # Alignments & formatting based on columns
                if c == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c == 2:
                    val_str = str(cell.value or "")
                    if "STAYED AT" in val_str:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.font = data_bold_font
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c in [3, 4, 5, 6]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c == 7:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c == 8:
                    if journey_type == "NORMAL":
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c == 9:
                    if journey_type == "NORMAL":
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0'
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif c == 10:
                    if journey_type == "NORMAL":
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0'
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0'
                elif c == 11:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0'

        # Row height auto-adjustment for data rows in pairs
        num_pairs = required_rows // 2
        for i in range(num_pairs):
            r1 = 9 + 2 * i
            r2 = 10 + 2 * i
            
            if journey_type == "NORMAL":
                obj_width = ws.column_dimensions['H'].width or 45
                obj_text = ws.cell(row=r1, column=8).value or ""
                needed_height = calculate_needed_row_height(obj_text, obj_width, font_size=9.5)
            else:
                obj_width = ws.column_dimensions['I'].width or 45
                obj_text = ws.cell(row=r1, column=9).value or ""
                needed_height = calculate_needed_row_height(obj_text, obj_width, font_size=9.5)
                
                # Check for stay details in training sheet (merged B-F, columns 2-6)
                stay_cell = ws.cell(row=r1, column=2)
                if is_merged_across_columns(ws, r1, 2):
                    combined_width = sum(ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width or 12 for col_idx in range(2, 7))
                    stay_text = stay_cell.value or ""
                    stay_height = calculate_needed_row_height(stay_text, combined_width, font_size=9.5, bold=True)
                    needed_height = max(needed_height, stay_height)
                    
            pair_height = max(44, needed_height)
            row_height = pair_height / 2
            ws.row_dimensions[r1].height = row_height
            ws.row_dimensions[r2].height = row_height

        # 4. Format Total Row
        ws.row_dimensions[total_row].height = 26
        
        for c in range(1, max_col + 1):
            cell = ws.cell(row=total_row, column=c)
            cell.fill = total_fill
            cell.border = total_border
            if cell.value:
                cell.font = total_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        # 5. Format Bottom Certification and Signature block
        cert_start = total_row + 1
        for r in range(cert_start, ws.max_row + 1):
            row_has_text = any(ws.cell(row=r, column=c).value is not None for c in range(1, max_col + 1))
            if row_has_text:
                has_signature_keywords = any(
                    isinstance(ws.cell(row=r, column=c).value, str) and 
                    any(k in ws.cell(row=r, column=c).value for k in ["Countersigned", "Signature of Officer", "Controlling Officer", "Head of Office"])
                    for c in range(1, max_col + 1)
                )
                has_cert_keywords = any(
                    isinstance(ws.cell(row=r, column=c).value, str) and 
                    "I hereby certify that" in ws.cell(row=r, column=c).value
                    for c in range(1, max_col + 1)
                )
                
                if has_signature_keywords:
                    ws.row_dimensions[r].height = 28
                elif has_cert_keywords:
                    ws.row_dimensions[r].height = 24
                else:
                    ws.row_dimensions[r].height = 18
            else:
                ws.row_dimensions[r].height = 10

            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value and isinstance(cell.value, str):
                    if "I hereby certify that" in cell.value:
                        cell.value = cell.value.strip()  # Remove leading spaces
                        cell.font = Font(name="Segoe UI", size=9, italic=True, color="475569")
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    elif any(k in cell.value for k in ["Countersigned", "Signature of Officer", "Controlling Officer", "Head of Office"]):
                        cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="1B365D")
                        if "Signature of Officer" in cell.value:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif "Note:" in cell.value:
                        cell.font = Font(name="Segoe UI", size=8, italic=True, color="64748B")
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Ensure grid lines are shown
        ws.views.sheetView[0].showGridLines = True

    conn = get_db()
    bill_row = conn.execute("""
        SELECT b.*, e.name as emp_name, e.pf_number, e.designation, e.level, e.joining_date, s.section_code
        FROM ta_bills b
        JOIN employees e ON b.emp_id = e.emp_id
        LEFT JOIN sections s ON e.primary_section_id = s.id
        WHERE b.id = ?
    """, (id,)).fetchone()
    
    if not bill_row:
        conn.close()
        raise HTTPException(status_code=404, detail="TA Bill not found")
        
    bill = dict(bill_row)
    entries_rows = conn.execute("SELECT * FROM ta_entries WHERE bill_id = ? ORDER BY id ASC", (id,)).fetchall()
    entries = [dict(r) for r in entries_rows]
    conn.close()
    
    if not os.path.exists(TA_TEMPLATE_PATH):
        raise HTTPException(status_code=500, detail="TA template Excel file not found on server.")
        
    try:
        wb = openpyxl.load_workbook(TA_TEMPLATE_PATH)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load TA template: {e}")
        
    journey_type = bill.get("journey_type", "NORMAL")
    sheet_to_keep = "format" if journey_type == "NORMAL" else "SUBRATA"
    
    if sheet_to_keep not in wb.sheetnames:
        raise HTTPException(status_code=500, detail=f"Template sheet '{sheet_to_keep}' not found in workbook.")
        
    ws = wb[sheet_to_keep]
    
    limit_row = 27 if journey_type == "NORMAL" else 19
    # Fix any invalid crossing merged ranges or template merged cells in the data area (row 9 onwards)
    for r in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = r.bounds
        if max_row >= 9 and min_row <= limit_row:
            try:
                ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
            except Exception:
                pass
            if min_row <= 8:
                new_max_row = min(8, max_row)
                try:
                    ws.merge_cells(start_row=min_row, start_column=min_col, end_row=new_max_row, end_column=max_col)
                except Exception:
                    pass
    
    emp_name = (bill.get("emp_name", "") or "").strip()
    if not emp_name:
        emp_name = "[Enter Employee Name]"
        
    pf_number = (bill.get("pf_number", "") or "").strip()
    if not pf_number:
        pf_number = "[Enter PF Number]"
        
    designation = (bill.get("designation", "") or "").strip()
    if not designation:
        designation = "[Enter Designation]"
        
    level = bill.get("level", 1) or 1
    basic_pay = bill.get("basic_pay", 0) or 0
    basic_pay_display = f"Rs. {basic_pay}" if basic_pay > 0 else "[Enter Basic Pay]"
    
    joining_date = (bill.get("joining_date", "") or "").strip()
    month_year = bill.get("month_year", "") or ""
    
    book_no = (bill.get("book_no", "") or "").strip()
    if not book_no:
        book_no = "[Enter Book Number]"
        
    page_no = (bill.get("page_no", "") or "").strip()
    if not page_no:
        page_no = "[Enter Page Number]"
        
    serial_no_from = (bill.get("serial_no_from", "") or "").strip()
    if not serial_no_from:
        serial_no_from = "[Start Serial]"
        
    serial_no_to = (bill.get("serial_no_to", "") or "").strip()
    if not serial_no_to:
        serial_no_to = "[End Serial]"
        
    bill_unit = (bill.get("bill_unit", "") or "").strip()
    if not bill_unit:
        bill_unit = "[Enter Bill Unit]"
        
    section_code = bill.get("section_code", "") or "KKVS"

    # Helpers to dynamically unmerge and format premium header blocks
    def write_meta_box(ws, start_row, start_col, end_row, end_col, label, value, bold_val=False, alignment="center"):
        for r_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = r_range.bounds
            if (start_row <= max_row and end_row >= min_row) and (start_col <= max_col and end_col >= min_col):
                try:
                    ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                except Exception:
                    pass
        thin_border = Border(
            left=Side(border_style="thin", color="CBD5E1"),
            right=Side(border_style="thin", color="CBD5E1"),
            top=Side(border_style="thin", color="CBD5E1"),
            bottom=Side(border_style="thin", color="CBD5E1")
        )
        bg_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        font_color = "1B365D" if bold_val else "333333"
        cell_font = Font(name="Segoe UI", size=9.5, bold=bold_val, color=font_color)
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.fill = bg_fill
                cell.border = thin_border
                cell.font = cell_font
                cell.alignment = Alignment(horizontal=alignment, vertical="center", wrap_text=True)
        top_left_cell = ws.cell(row=start_row, column=start_col)
        top_left_cell.value = f"{label}: {value}" if label else value
        if start_row != end_row or start_col != end_col:
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

    def write_title_box(ws, start_row, start_col, end_row, end_col, title):
        for r_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = r_range.bounds
            if (start_row <= max_row and end_row >= min_row) and (start_col <= max_col and end_col >= min_col):
                try:
                    ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                except Exception:
                    pass
        title_font = Font(name="Segoe UI", size=13, bold=True, color="1B365D")
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.font = title_font
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=start_row, column=start_col).value = title
        if start_row != end_row or start_col != end_col:
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

    try:
        dt = datetime.strptime(month_year, "%Y-%m")
        month_year_display = dt.strftime("%B  %Y").upper()
    except Exception:
        month_year_display = month_year.upper()

    joining_date_display = ""
    if joining_date:
        try:
            dt_ap = datetime.strptime(joining_date, "%Y-%m-%d")
            joining_date_display = dt_ap.strftime("%d.%m.%Y")
        except Exception:
            joining_date_display = joining_date
    else:
        joining_date_display = "[Enter Joining Date]"

    # Build the header grid fields cleanly
    if journey_type == "NORMAL":
        # Row 1
        write_meta_box(ws, 1, 1, 1, 4, "", f"TA Movement Register Book No. {book_no} at {section_code}", bold_val=True, alignment="left")
        write_meta_box(ws, 1, 9, 1, 10, "", f"P.F. No. {pf_number}", bold_val=True, alignment="right")
        # Row 2
        write_meta_box(ws, 2, 1, 2, 2, "", f"Page No. {page_no}", alignment="left")
        write_meta_box(ws, 2, 3, 2, 4, "", f"Serial No. from {serial_no_from} to {serial_no_to}", alignment="center")
        write_meta_box(ws, 2, 9, 2, 10, "", f"B.U. No. {bill_unit}", alignment="right")
        write_title_box(ws, 1, 5, 2, 8, "TRAVELLING ALLOWANCE JOURNAL")
        # Row 4
        write_meta_box(ws, 4, 1, 4, 2, "Department", "S&T", alignment="left")
        write_meta_box(ws, 4, 3, 4, 4, "Division Headquarters at", section_code, alignment="center")
        write_meta_box(ws, 4, 5, 4, 8, "Journal of duties performed by", f"Sri {emp_name}", bold_val=True, alignment="left")
        write_meta_box(ws, 4, 9, 4, 10, "For which allowance is claimed", month_year_display, alignment="right")
        # Row 5
        write_meta_box(ws, 5, 1, 5, 1, "is claimed", "TA", alignment="left")
        write_meta_box(ws, 5, 2, 5, 3, "Designation", designation, alignment="left")
        write_meta_box(ws, 5, 4, 5, 5, "Pay", basic_pay_display, alignment="center")
        write_meta_box(ws, 5, 6, 5, 6, "Level", level, alignment="center")
        write_meta_box(ws, 5, 7, 5, 8, "Date of appointment", joining_date_display, alignment="center")
        write_meta_box(ws, 5, 9, 5, 10, "Rule by which governed", "SRTA", alignment="right")
    else:
        # Row 1
        write_meta_box(ws, 1, 1, 1, 4, "", f"TA Movement Register Book No. {book_no} at {section_code}", bold_val=True, alignment="left")
        write_meta_box(ws, 1, 10, 1, 11, "", f"P.F. No. {pf_number}", bold_val=True, alignment="right")
        # Row 2
        write_meta_box(ws, 2, 1, 2, 2, "", f"Page No. {page_no}", alignment="left")
        write_meta_box(ws, 2, 3, 2, 4, "", f"Serial No. from {serial_no_from} to {serial_no_to}", alignment="center")
        write_meta_box(ws, 2, 10, 2, 11, "", f"B.U. No. {bill_unit}", alignment="right")
        write_title_box(ws, 1, 5, 2, 9, "TRAVELLING ALLOWANCE JOURNAL")
        # Row 4
        write_meta_box(ws, 4, 1, 4, 2, "Department", "S&T", alignment="left")
        write_meta_box(ws, 4, 3, 4, 4, "Division Headquarters at", section_code, alignment="center")
        write_meta_box(ws, 4, 5, 4, 9, "Journal of duties performed by", f"Sri {emp_name}", bold_val=True, alignment="left")
        write_meta_box(ws, 4, 10, 4, 11, "For which allowance is claimed", month_year_display, alignment="right")
        # Row 5
        write_meta_box(ws, 5, 1, 5, 1, "is claimed", "TA", alignment="left")
        write_meta_box(ws, 5, 2, 5, 3, "Designation", designation, alignment="left")
        write_meta_box(ws, 5, 4, 5, 5, "Pay", basic_pay_display, alignment="center")
        write_meta_box(ws, 5, 6, 5, 6, "Level", level, alignment="center")
        write_meta_box(ws, 5, 7, 5, 9, "Date of appointment", joining_date_display, alignment="center")
        write_meta_box(ws, 5, 10, 5, 11, "Rule by which governed", "SRTA", alignment="right")

    # Write data
    total_amount = 0
    if journey_type == "NORMAL":
        num_pairs = (len(entries) + 1) // 2
        required_rows = num_pairs * 2
        
        if required_rows > 18:
            insert_rows_clean(ws, 27, required_rows - 18)
            for r in range(27, 27 + required_rows - 18):
                if r % 2 == 1:
                    copy_row_style(ws, 9, r)
                else:
                    copy_row_style(ws, 10, r)
        elif required_rows < 18:
            delete_rows_clean(ws, 9 + required_rows, 18 - required_rows)
            
        entry_start = 9
        entry_end = 8 + required_rows
        for r in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = r.bounds
            if min_row >= entry_start and max_row <= entry_end:
                ws.merged_cells.ranges.remove(r)
                
        for i in range(num_pairs):
            r1_idx = 9 + 2 * i
            r2_idx = 10 + 2 * i
            
            leg_out = entries[2 * i]
            leg_in = entries[2 * i + 1] if (2 * i + 1) < len(entries) else None
            
            date_val = leg_out.get("entry_date", "")
            try:
                dt_d = datetime.strptime(date_val, "%Y-%m-%d")
                date_display = dt_d.strftime("%d.%m.%y")
            except Exception:
                date_display = date_val
                
            ws.cell(row=r1_idx, column=1).value = date_display
            ws.cell(row=r1_idx, column=2).value = leg_out.get("train_no", "")
            
            t_left = leg_out.get("time_left", "")
            if t_left and not any(h in t_left.lower() for h in ("hrs", "hr")):
                t_left = f"{t_left} hrs"
            ws.cell(row=r1_idx, column=3).value = t_left
            
            t_arr = leg_out.get("time_arrived", "")
            if t_arr and not any(h in t_arr.lower() for h in ("hrs", "hr")):
                t_arr = f"{t_arr} hrs"
            ws.cell(row=r1_idx, column=4).value = t_arr
            
            ws.cell(row=r1_idx, column=5).value = leg_out.get("station_from", "")
            ws.cell(row=r1_idx, column=6).value = leg_out.get("station_to", "")
            
            days_nights_val = leg_out.get("days_nights", "")
            try:
                days_nights_display = float(days_nights_val) if days_nights_val else None
            except Exception:
                days_nights_display = days_nights_val
            ws.cell(row=r1_idx, column=7).value = days_nights_display
            
            ws.cell(row=r1_idx, column=8).value = leg_out.get("object_journey", "")
            ws.cell(row=r1_idx, column=9).value = leg_out.get("rate", 0)
            ws.cell(row=r1_idx, column=10).value = leg_out.get("amount", 0)
            total_amount += leg_out.get("amount", 0)
            
            if leg_in:
                ws.cell(row=r2_idx, column=2).value = leg_in.get("train_no", "")
                
                t_left_in = leg_in.get("time_left", "")
                if t_left_in and not any(h in t_left_in.lower() for h in ("hrs", "hr")):
                    t_left_in = f"{t_left_in} hrs"
                ws.cell(row=r2_idx, column=3).value = t_left_in
                
                t_arr_in = leg_in.get("time_arrived", "")
                if t_arr_in and not any(h in t_arr_in.lower() for h in ("hrs", "hr")):
                    t_arr_in = f"{t_arr_in} hrs"
                ws.cell(row=r2_idx, column=4).value = t_arr_in
                
                ws.cell(row=r2_idx, column=5).value = leg_in.get("station_from", "")
                ws.cell(row=r2_idx, column=6).value = leg_in.get("station_to", "")
            else:
                for c in range(2, 7):
                    ws.cell(row=r2_idx, column=c).value = ""
                    
            ws.merge_cells(start_row=r1_idx, start_column=1, end_row=r2_idx, end_column=1)
            ws.merge_cells(start_row=r1_idx, start_column=7, end_row=r2_idx, end_column=7)
            ws.merge_cells(start_row=r1_idx, start_column=8, end_row=r2_idx, end_column=8)
            ws.merge_cells(start_row=r1_idx, start_column=9, end_row=r2_idx, end_column=9)
            ws.merge_cells(start_row=r1_idx, start_column=10, end_row=r2_idx, end_column=10)
            
        total_row_idx = 9 + required_rows
        ws.cell(row=total_row_idx, column=8).value = f"Total :Rupees {num_to_words(total_amount)} only                     Rs.              {total_amount}"
        ws.merge_cells(start_row=total_row_idx, start_column=8, end_row=total_row_idx, end_column=10)

    else:
        N = len(entries)
        required_rows = N * 2
        
        if required_rows > 10:
            insert_rows_clean(ws, 19, required_rows - 10)
            for r in range(19, 19 + required_rows - 10):
                if r % 2 == 1:
                    copy_row_style(ws, 9, r)
                else:
                    copy_row_style(ws, 10, r)
        elif required_rows < 10:
            delete_rows_clean(ws, 9 + required_rows, 10 - required_rows)
            
        entry_start = 9
        entry_end = 8 + required_rows
        for r in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = r.bounds
            if min_row >= entry_start and max_row <= entry_end:
                ws.merged_cells.ranges.remove(r)
                
        for i in range(N):
            r1_idx = 9 + 2 * i
            r2_idx = 10 + 2 * i
            entry = entries[i]
            is_stay = entry.get("is_stay", 0)
            
            date_val = entry.get("entry_date", "")
            try:
                dt_d = datetime.strptime(date_val, "%Y-%m-%d")
                date_display = dt_d.strftime("%d.%m.%y")
            except Exception:
                date_display = date_val
                
            ws.cell(row=r1_idx, column=1).value = date_display
            ws.merge_cells(start_row=r1_idx, start_column=1, end_row=r2_idx, end_column=1)
            
            days_nights_val = entry.get("days_nights", "")
            ws.cell(row=r1_idx, column=8).value = days_nights_val
            ws.merge_cells(start_row=r1_idx, start_column=8, end_row=r2_idx, end_column=8)
            
            ws.cell(row=r1_idx, column=9).value = entry.get("object_journey", "")
            ws.merge_cells(start_row=r1_idx, start_column=9, end_row=r2_idx, end_column=9)
            
            ws.cell(row=r1_idx, column=10).value = entry.get("rate", 0)
            ws.merge_cells(start_row=r1_idx, start_column=10, end_row=r2_idx, end_column=10)
            
            ws.cell(row=r1_idx, column=11).value = entry.get("amount", 0)
            ws.merge_cells(start_row=r1_idx, start_column=11, end_row=r2_idx, end_column=11)
            total_amount += entry.get("amount", 0)
            
            ws.cell(row=r1_idx, column=7).value = None
            ws.merge_cells(start_row=r1_idx, start_column=7, end_row=r2_idx, end_column=7)
            
            if is_stay:
                stay_text = entry.get("stay_details") or entry.get("train_no") or ""
                ws.cell(row=r1_idx, column=2).value = stay_text
                ws.merge_cells(start_row=r1_idx, start_column=2, end_row=r2_idx, end_column=6)
            else:
                ws.cell(row=r1_idx, column=2).value = entry.get("train_no", "")
                
                t_left = entry.get("time_left", "")
                if t_left and not any(h in t_left.lower() for h in ("hrs", "hr")):
                    t_left = f"{t_left} hrs"
                ws.cell(row=r1_idx, column=3).value = t_left
                
                t_arr = entry.get("time_arrived", "")
                if t_arr and not any(h in t_arr.lower() for h in ("hrs", "hr")):
                    t_arr = f"{t_arr} hrs"
                ws.cell(row=r1_idx, column=4).value = t_arr
                
                ws.cell(row=r1_idx, column=5).value = entry.get("station_from", "")
                ws.cell(row=r1_idx, column=6).value = entry.get("station_to", "")
                
                ws.merge_cells(start_row=r1_idx, start_column=2, end_row=r2_idx, end_column=2)
                ws.merge_cells(start_row=r1_idx, start_column=3, end_row=r2_idx, end_column=3)
                ws.merge_cells(start_row=r1_idx, start_column=4, end_row=r2_idx, end_column=4)
                ws.merge_cells(start_row=r1_idx, start_column=5, end_row=r2_idx, end_column=5)
                ws.merge_cells(start_row=r1_idx, start_column=6, end_row=r2_idx, end_column=6)
                
        total_row_idx = 9 + required_rows
        ws.cell(row=total_row_idx, column=9).value = f"Total :Rupees {num_to_words(total_amount)} only                     Rs.              {total_amount}"
        ws.merge_cells(start_row=total_row_idx, start_column=9, end_row=total_row_idx, end_column=11)

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=r, column=c).value
            if cell_val and isinstance(cell_val, str) and "I hereby certify that above mentioned" in cell_val:
                new_text = re.sub(
                    r"(above mentioned\s+)(.*?)(absent on duty)", 
                    r"\g<1>  Sri " + emp_name + r"  \g<3>", 
                    cell_val
                )
                ws.cell(row=r, column=c).value = new_text

    # Apply dynamic premium layout styling
    apply_premium_styling(ws, journey_type, required_rows)

    sheet_name_clean = emp_name[:30].strip()
    ws.title = sheet_name_clean
    for s in list(wb.sheetnames):
        if s != sheet_name_clean:
            wb.remove(wb[s])
            
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"TA_Bill_{emp_name.replace(' ', '_')}_{month_year}.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Resolve frontend/out directory
if getattr(sys, 'frozen', False):
    frontend_out_dir = os.path.join(sys._MEIPASS, "out")
else:
    frontend_out_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend", "out"))

# Mount _next folder for static resources
if os.path.exists(os.path.join(frontend_out_dir, "_next")):
    app.mount("/_next", StaticFiles(directory=os.path.join(frontend_out_dir, "_next")), name="next-assets")

# Route to serve other static files or HTML clean URLs
@app.get("/{path_name:path}")
async def serve_frontend(path_name: str):
    # Strip leading/trailing slashes
    clean_path = path_name.strip("/")
    
    if not clean_path:
        index_path = os.path.join(frontend_out_dir, "index.html")
        if os.path.isfile(index_path):
            return FileResponse(index_path)
            
    # 1. Try to serve exact file from out directory (e.g. favicon.ico, app_logo.png, etc.)
    file_path = os.path.join(frontend_out_dir, clean_path)
    if os.path.isfile(file_path):
        return FileResponse(file_path)
        
    # 2. Try clean routing (e.g. clean_path = "employees" -> employees.html)
    html_path = os.path.join(frontend_out_dir, f"{clean_path}.html")
    if os.path.isfile(html_path):
        return FileResponse(html_path)
        
    # 3. Fallback to 404 or index.html for client-side routing
    four_oh_four = os.path.join(frontend_out_dir, "404.html")
    if os.path.isfile(four_oh_four):
        return FileResponse(four_oh_four)
    return FileResponse(os.path.join(frontend_out_dir, "index.html"))

if __name__ == "__main__":
    import uvicorn
    is_frozen = getattr(sys, 'frozen', False)
    if is_frozen:
        # Pass the app object directly and disable reload to avoid multiprocessing issues in PyInstaller
        uvicorn.run(app, host="127.0.0.1", port=8000)
    else:
        uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)


